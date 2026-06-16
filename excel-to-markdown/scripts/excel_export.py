"""
Excel 三通道导出工具 — 为 Agent 分析优化

输出结构:
    output_dir/
        ├── {SheetName}/
        │   ├── screenshot.png          # 高清截图（保留图片、颜色、排版）
        │   ├── screenshot_part2.png    # 大 Sheet 分片（如有）
        │   └── data.md                 # 精确结构化文本 + 富文本标注 + 表格检测

前提条件:
    - Windows + Microsoft Excel
    - pip install pywin32 openpyxl Pillow beautifulsoup4 cssutils PyMuPDF

用法:
    python excel_export.py <input.xlsx> [output_dir]
"""

import sys
import os

# Windows: stdout/stderr 统一使用 UTF-8，避免中文乱码
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import re
import time
import ctypes
import ctypes.wintypes
import logging
import hashlib
import json
import shutil
from pathlib import Path
from typing import Optional, List, Tuple, Dict
from dataclasses import dataclass, field

import win32com.client
import pythoncom
import cssutils
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import ImageGrab, Image
from bs4 import BeautifulSoup

# cssutils 默认日志太吵，静音
cssutils.log.setLevel(logging.CRITICAL)


# ═══════════════════════════════════════════════════════════
#  结构化数据模型
# ═══════════════════════════════════════════════════════════

@dataclass
class CellStyle:
    """单元格样式信息（统一数据模型，三通道数据在此汇合）"""
    bold: bool = False
    italic: bool = False
    strikethrough: bool = False
    underline: bool = False
    color: Optional[str] = None       # 文字颜色 (hex, e.g. "#FF0000")
    bg_color: Optional[str] = None    # 背景色 (hex, e.g. "#FFFF00")
    font_size: Optional[float] = None
    has_border: bool = False           # 是否有显式边框（用于表格区域检测）
    text_align: Optional[str] = None


@dataclass
class CellData:
    """单元格数据（统一数据模型）"""
    row: int
    col: int
    text: str = ""
    style: CellStyle = field(default_factory=CellStyle)
    colspan: int = 1
    rowspan: int = 1
    is_merged_slave: bool = False      # 被合并覆盖的从属单元格
    images: list = field(default_factory=list)
    html_content: str = ""             # 原始 HTML 内容（用于提取内联混合格式）


# ═══════════════════════════════════════════════════════════
#  颜色智能过滤
# ═══════════════════════════════════════════════════════════

def _expand_short_hex(hex_color: str) -> str:
    """将 3 位缩写 hex (#RGB) 展开为 6 位 (#RRGGBB)"""
    if hex_color and len(hex_color) == 4 and hex_color.startswith('#'):
        r, g, b = hex_color[1], hex_color[2], hex_color[3]
        return f'#{r}{r}{g}{g}{b}{b}'
    return hex_color


def _is_near_white_or_grey(hex_color: str) -> bool:
    """
    判断颜色是否接近白色/浅灰色（在 Markdown 白底上不可见）。
    RGB 各通道均 >= 180 视为不可见。
    支持 #RGB 缩写和 #RRGGBB 两种格式。
    """
    if not hex_color:
        return False
    hex_color = _expand_short_hex(hex_color)
    if len(hex_color) != 7:
        return False
    try:
        r = int(hex_color[1:3], 16)
        g = int(hex_color[3:5], 16)
        b = int(hex_color[5:7], 16)
        return r >= 180 and g >= 180 and b >= 180
    except ValueError:
        return False


_NAMED_COLORS = {
    'red': '#FF0000', 'blue': '#0000FF', 'green': '#008000',
    'yellow': '#FFFF00', 'purple': '#800080', 'orange': '#FFA500',
    'gray': '#808080', 'grey': '#808080', 'pink': '#FFC0CB',
    'cyan': '#00FFFF', 'lime': '#00FF00',
}


def _normalize_color(color_str: str) -> Optional[str]:
    """将颜色值规范化为 #RRGGBB 格式，过滤 Markdown 白底上不可见的颜色"""
    if not color_str:
        return None
    color_str = color_str.strip().lower()
    if color_str in ('black', 'windowtext', 'white', 'inherit', ''):
        return None

    if color_str.startswith('#'):
        hex_color = _expand_short_hex(color_str.upper())
        # 黑色是默认文字颜色，无需标注
        if hex_color in ('#000000',):
            return None
        return hex_color if not _is_near_white_or_grey(hex_color) else None

    if color_str in _NAMED_COLORS:
        return _NAMED_COLORS[color_str]

    # rgb(r,g,b)
    m = re.match(r'rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', color_str)
    if m:
        hex_color = f'#{int(m.group(1)):02X}{int(m.group(2)):02X}{int(m.group(3)):02X}'
        return hex_color if not _is_near_white_or_grey(hex_color) else None

    return None


def _normalize_bg_color(color_str: str) -> Optional[str]:
    """将背景色规范化为 #RRGGBB，过滤白色/浅灰等在 Markdown 中无意义的背景色"""
    if not color_str:
        return None
    color_str = color_str.strip().lower()
    if color_str in ('white', 'inherit', 'transparent', 'window', ''):
        return None

    if color_str in _NAMED_COLORS:
        return _NAMED_COLORS[color_str]

    if color_str.startswith('#'):
        hex_color = _expand_short_hex(color_str.upper())
        return None if _is_near_white_or_grey(hex_color) else hex_color

    m = re.match(r'rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', color_str)
    if m:
        hex_color = f'#{int(m.group(1)):02X}{int(m.group(2)):02X}{int(m.group(3)):02X}'
        return None if _is_near_white_or_grey(hex_color) else hex_color

    return None


# ═══════════════════════════════════════════════════════════
#  CSS 双策略解析（用于 HTML 通道）
# ═══════════════════════════════════════════════════════════

def _parse_css_classes(css_text: str) -> Dict[str, CellStyle]:
    """
    解析 Excel 导出 HTML 的 CSS，建立 class → CellStyle 映射。
    主策略: cssutils 库解析；备选: 正则表达式（cssutils 失败时回退）。
    """
    try:
        sheet = cssutils.parseString(css_text, validate=False)
    except Exception:
        return _parse_css_regex(css_text)

    class_styles = {}
    for rule in sheet:
        if rule.type != rule.STYLE_RULE:
            continue
        selector = rule.selectorText.strip()
        if not selector.startswith('.'):
            continue

        cls_name = selector[1:]
        style = CellStyle()
        s = rule.style

        # 文字颜色
        color = s.getPropertyValue('color')
        if color and color not in ('black', 'windowtext', 'white', 'inherit', ''):
            style.color = _normalize_color(color)

        # 加粗
        fw = s.getPropertyValue('font-weight')
        if fw and (fw == 'bold' or (fw.isdigit() and int(fw) >= 700)):
            style.bold = True

        # 斜体
        fs = s.getPropertyValue('font-style')
        if fs == 'italic':
            style.italic = True

        # 删除线 / 下划线
        td = s.getPropertyValue('text-decoration')
        if td:
            if 'line-through' in td:
                style.strikethrough = True
            if 'underline' in td:
                style.underline = True

        # 字号
        fsize = s.getPropertyValue('font-size')
        if fsize:
            m = re.search(r'([\d.]+)', fsize)
            if m:
                style.font_size = float(m.group(1))

        # 边框（任何方向有边框就算）
        for prop in ['border', 'border-top', 'border-right', 'border-bottom', 'border-left']:
            val = s.getPropertyValue(prop)
            if val and ('solid' in val or 'dashed' in val or 'dotted' in val or 'double' in val) and 'none' not in val:
                style.has_border = True
                break

        # 背景色
        bg = s.getPropertyValue('background')
        if bg and bg not in ('white', 'inherit', '', 'transparent', 'window'):
            style.bg_color = _normalize_bg_color(bg)

        # 文字对齐
        ta = s.getPropertyValue('text-align')
        if ta and ta not in ('general', ''):
            style.text_align = ta

        class_styles[cls_name] = style

    return class_styles


def _parse_css_regex(css_text: str) -> Dict[str, CellStyle]:
    """正则方式解析 CSS（cssutils 失败时的备选策略）"""
    class_styles = {}
    blocks = re.findall(r'\.(\w+)\s*\{([^}]*)\}', css_text, re.DOTALL)
    for cls_name, props in blocks:
        style = CellStyle()

        # color
        m = re.search(r'(?<![a-z-])color\s*:\s*([^;]+)', props)
        if m:
            c = m.group(1).strip()
            if c not in ('black', 'windowtext', 'white', 'inherit'):
                style.color = _normalize_color(c)

        # font-weight
        m = re.search(r'font-weight\s*:\s*(\w+)', props)
        if m and (m.group(1) == 'bold' or (m.group(1).isdigit() and int(m.group(1)) >= 700)):
            style.bold = True

        # font-style
        if re.search(r'font-style\s*:\s*italic', props):
            style.italic = True

        # text-decoration
        m = re.search(r'text-decoration\s*:\s*([^;]+)', props)
        if m:
            if 'line-through' in m.group(1):
                style.strikethrough = True
            if 'underline' in m.group(1):
                style.underline = True

        # font-size
        m = re.search(r'font-size\s*:\s*([\d.]+)', props)
        if m:
            style.font_size = float(m.group(1))

        # background
        m = re.search(r'(?<![a-z-])background\s*:\s*([^;]+)', props)
        if m:
            bg = m.group(1).strip()
            if bg not in ('white', 'inherit', 'transparent', 'window'):
                style.bg_color = _normalize_bg_color(bg)

        # border（增强：支持 dashed/dotted/double）
        for prop in ['border', 'border-top', 'border-right', 'border-bottom', 'border-left']:
            pattern = rf'{prop}\s*:\s*([^;]+)'
            bm = re.search(pattern, props)
            if bm:
                bval = bm.group(1)
                if ('solid' in bval or 'dashed' in bval or 'dotted' in bval or 'double' in bval) and 'none' not in bval:
                    style.has_border = True
                    break

        class_styles[cls_name] = style

    return class_styles


def _merge_cell_style(target: CellStyle, source: CellStyle):
    """将 source 样式合并到 target（source 的非默认值覆盖 target）"""
    if source.bold: target.bold = True
    if source.italic: target.italic = True
    if source.strikethrough: target.strikethrough = True
    if source.underline: target.underline = True
    if source.color: target.color = source.color
    if source.bg_color: target.bg_color = source.bg_color
    if source.font_size: target.font_size = source.font_size
    if source.has_border: target.has_border = True
    if source.text_align: target.text_align = source.text_align


# ═══════════════════════════════════════════════════════════
#  HTML 通道：解析 Excel COM 导出的 HTML
# ═══════════════════════════════════════════════════════════

def _extract_cell_text(td) -> str:
    """从 <td> 提取纯文本，将 <br> 转为换行"""
    for br in td.find_all('br'):
        br.replace_with('\n')
    return td.get_text()


def _read_html_auto_encoding(file_path: str) -> str:
    """自动检测编码读取 HTML 文件（优先 UTF-8，回退 GB2312/GBK）"""
    # 先尝试 UTF-8
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        # 简单验证：UTF-8 解码成功且不含大量替换字符
        if '\ufffd' not in content[:2000]:
            return content
    except (UnicodeDecodeError, UnicodeError):
        pass

    # 尝试 GBK（GB2312 的超集）
    try:
        with open(file_path, 'r', encoding='gbk', errors='replace') as f:
            return f.read()
    except Exception:
        pass

    # 最终回退：latin-1（不会报错，但可能乱码）
    with open(file_path, 'r', encoding='latin-1') as f:
        return f.read()


def _clean_excel_html(html: str) -> str:
    """
    预清洗 Excel 导出的 HTML，去除大量微软特有冗余内容。
    可将 180MB 的 HTML 缩减到原来的 10-30%，大幅加速后续 BeautifulSoup 解析。
    """
    original_len = len(html)

    # 1. 去除微软条件注释 <!--[if ...]>...<![endif]-->（VML、Office 特有标记）
    #    这些可能占 HTML 体积的 30-60%
    html = re.sub(r'<!--\[if[^]]*\]>.*?<!\[endif\]-->', '', html, flags=re.DOTALL)
    html = re.sub(r'<!--\[if[^]]*\]>', '', html, flags=re.DOTALL)
    html = re.sub(r'<!\[endif\]-->', '', html)

    # 2. 去除 <style> 标签中的冗余内容（CSS 已单独解析）
    #    注意：sheet HTML 里的 <style> 一般是重复的，CSS 已从 .css 文件解析
    # （保留 <style> 以防某些情况下内联 CSS 有用）

    # 3. 去除 XML 命名空间声明和 Office 特有标签
    html = re.sub(r'<o:[^>]*>.*?</o:[^>]*>', '', html, flags=re.DOTALL)
    html = re.sub(r'<o:[^>]*/>', '', html)
    html = re.sub(r'<x:[^>]*>.*?</x:[^>]*>', '', html, flags=re.DOTALL)
    html = re.sub(r'<x:[^>]*/>', '', html)
    html = re.sub(r'<v:[^>]*>.*?</v:[^>]*>', '', html, flags=re.DOTALL)
    html = re.sub(r'<v:[^>]*/>', '', html)

    # 4. 去除 HTML 注释（普通注释）
    html = re.sub(r'<!--(?!\[).*?-->', '', html, flags=re.DOTALL)

    # 5. 去除 <col> 标签中的 style 属性（只保留结构）
    html = re.sub(r'(<col\b)[^>]*(>|/>)', r'\1\2', html)

    cleaned_len = len(html)
    if original_len > 0:
        ratio = cleaned_len / original_len * 100
        print(f"    [预清洗] HTML {original_len // 1024}KB → {cleaned_len // 1024}KB ({ratio:.0f}%)")

    return html


# 预编译正则（模块级，避免每次调用重新编译）
_RE_TABLE = re.compile(r'<table[^>]*>(.*)</table>', re.DOTALL)
_RE_TR = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL)
_RE_TD = re.compile(r'<t[dh]\b([^>]*)>(.*?)</t[dh]>', re.DOTALL)
_RE_CLASS = re.compile(r'class="([^"]+)"')
_RE_COLSPAN = re.compile(r'colspan="(\d+)"')
_RE_ROWSPAN = re.compile(r'rowspan="(\d+)"')
_RE_TAG = re.compile(r'<[^>]+>')
_RE_BR = re.compile(r'<br\s*/?\s*>', re.IGNORECASE)
_RE_CSS_CLASS_TAG = re.compile(r'<(span|font)\s+class="([^"]+)"[^>]*>(.*?)</\1>', re.DOTALL)

# 内联格式标签检测关键字
_INLINE_FMT_KEYWORDS = ('<font', '<b>', '<b ', '<strong', '<s>', '<s ', '<del', '<i>', '<i ', '<em', '<u>', '<u ', '<span')


def _resolve_css_class_to_inline_tags(html_content: str, css_classes: Dict[str, CellStyle]) -> str:
    """
    将 HTML 中带 CSS class 的标签替换为等效的显式内联格式标签（<font color>/<b>/<i> 等）。

    Excel COM 导出 HTML 时，同一单元格内的混合格式（部分文字不同颜色/加粗）
    可能用以下方式表示：
    1. <span class="xl78">文字</span> — CSS class 定义样式
    2. <font class="font33">文字</font> — font 标签用 class 而非 color 属性

    此函数将它们展开为显式的 <font color="...">, <b>, <i> 等标签，
    使后续 _format_inline_html 能正确提取所有样式。
    """
    if not css_classes:
        return html_content

    def _replace_css_tag(match):
        tag_name = match.group(1)
        cls_names = match.group(2).split()
        inner = match.group(3)

        merged = CellStyle()
        has_css_match = False
        for cls in cls_names:
            if cls in css_classes:
                _merge_cell_style(merged, css_classes[cls])
                has_css_match = True

        if not has_css_match:
            return match.group(0)

        result = inner
        if merged.strikethrough:
            result = f'<s>{result}</s>'
        if merged.bold:
            result = f'<b>{result}</b>'
        if merged.italic:
            result = f'<i>{result}</i>'
        if merged.underline:
            result = f'<u>{result}</u>'
        if merged.color:
            result = f'<font color="{merged.color}">{result}</font>'

        return result

    return _RE_CSS_CLASS_TAG.sub(_replace_css_tag, html_content)


def _parse_sheet_html(html_path: str, css_classes: Dict[str, CellStyle]) -> List[List[CellData]]:
    """
    解析 Excel 导出的 sheet HTML，返回二维 CellData 数组。
    保留内联混合格式的 html_content，供 Markdown 生成时使用。

    使用正则而非 BeautifulSoup 解析，速度提升约 5 倍。
    Excel COM 导出的 HTML 结构高度规律（机器生成），正则解析安全可靠。
    """
    import html as html_module  # html.unescape

    raw_html = _read_html_auto_encoding(html_path)

    # 预清洗：去除微软特有冗余（VML/条件注释等），大幅减小解析量
    raw_html = _clean_excel_html(raw_html)

    # 提取 <table>...</table>
    table_match = _RE_TABLE.search(raw_html)
    if not table_match:
        return []
    table_html = table_match.group(1)

    # 提取所有 <tr>
    trs = _RE_TR.findall(table_html)

    rows_data = []
    rowspan_tracker = {}  # (row, col) -> CellData

    for row_idx, tr_content in enumerate(trs):
        row = []
        col_idx = 0

        for td_match in _RE_TD.finditer(tr_content):
            # 跳过被 rowspan 占位的列
            while (row_idx, col_idx) in rowspan_tracker:
                slave = CellData(row=row_idx, col=col_idx, is_merged_slave=True)
                row.append(slave)
                col_idx += 1

            attrs_str = td_match.group(1)
            content = td_match.group(2)

            # 解析属性
            cls_match = _RE_CLASS.search(attrs_str)
            classes = cls_match.group(1).split() if cls_match else []

            colspan_match = _RE_COLSPAN.search(attrs_str)
            colspan = int(colspan_match.group(1)) if colspan_match else 1

            rowspan_match = _RE_ROWSPAN.search(attrs_str)
            rowspan = int(rowspan_match.group(1)) if rowspan_match else 1

            # 合并 CSS class 的样式
            cell_style = CellStyle()
            for cls in classes:
                if cls in css_classes:
                    _merge_cell_style(cell_style, css_classes[cls])

            # 提取纯文本：<br> → \n，去掉所有标签，解码 HTML 实体
            text = _RE_BR.sub('\n', content)
            text = _RE_TAG.sub('', text)
            text = html_module.unescape(text)

            # 检测内联格式标签 → 保留 html_content
            html_content = ""
            has_inline = any(kw in content for kw in _INLINE_FMT_KEYWORDS)
            if has_inline:
                # 将 CSS class 标签展开为等效内联标签（<font color>/<b> 等）
                # 处理 <span class="xl78"> 和 <font class="font33"> 两种形式
                html_content = _resolve_css_class_to_inline_tags(content, css_classes)
                # 设置内联格式标志（在 CSS class 展开后重新检测）
                if '<s>' in html_content or '<s ' in html_content or '<del' in html_content:
                    cell_style.strikethrough = True
                if '<b>' in html_content or '<b ' in html_content or '<strong' in html_content:
                    cell_style.bold = True
                if '<i>' in html_content or '<i ' in html_content or '<em' in html_content:
                    cell_style.italic = True
                if '<u>' in html_content or '<u ' in html_content:
                    cell_style.underline = True

            cell = CellData(
                row=row_idx, col=col_idx,
                text=text, style=cell_style,
                colspan=colspan, rowspan=rowspan,
                html_content=html_content,
            )
            row.append(cell)

            # 注册 rowspan 占位
            if rowspan > 1:
                for r in range(row_idx + 1, row_idx + rowspan):
                    for c in range(col_idx, col_idx + colspan):
                        rowspan_tracker[(r, c)] = cell

            # 注册 colspan 占位
            for c in range(1, colspan):
                slave = CellData(row=row_idx, col=col_idx + c, is_merged_slave=True)
                row.append(slave)

            col_idx += colspan

        # 补充尾部的 rowspan 占位
        while (row_idx, col_idx) in rowspan_tracker:
            slave = CellData(row=row_idx, col=col_idx, is_merged_slave=True)
            row.append(slave)
            col_idx += 1

        rows_data.append(row)

    return rows_data


def _export_sheets_to_html(excel_app, wb_com, output_dir: Path,
                           sheet_names: List[str],
                           bounds_map: Optional[Dict[str, Tuple[int, int, int, int]]] = None) -> Dict[str, str]:
    """
    在已有 COM 会话中将指定 sheet 导出为 HTML。
    返回 {sheet_name: htm_file_path}。

    优化：当提供 bounds_map 时，在 Copy 后的临时 workbook 中删除有效内容范围之外
    的行/列，大幅减少导出 HTML 的体积（大 sheet 从 180MB 降至 <5MB）。
    """
    tmp_dir = output_dir / "_tmp_html"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    result = {}
    for i in range(1, wb_com.Sheets.Count + 1):
        ws = wb_com.Sheets(i)
        name = ws.Name
        if name not in sheet_names:
            continue

        safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in name)
        try:
            ws.Copy()
            new_wb = excel_app.ActiveWorkbook
            new_ws = new_wb.Sheets(1)

            # ── 裁剪超出 content_bounds 的行/列，减小 HTML 体积 ──
            bounds = bounds_map.get(name) if bounds_map else None
            if bounds:
                min_r, max_r, min_c, max_c = bounds
                # 加一点余量防止丢边界样式
                max_r_keep = min(max_r + 2, 2000)
                max_c_keep = min(max_c + 2, 200)

                # 删除右侧多余列（从右往左删，避免索引偏移）
                try:
                    total_cols = new_ws.UsedRange.Columns.Count
                    if total_cols > max_c_keep:
                        del_start_col = max_c_keep + 1
                        # 批量删除：选中范围一次性删除
                        col_range = new_ws.Range(
                            new_ws.Cells(1, del_start_col),
                            new_ws.Cells(1, total_cols)
                        ).EntireColumn
                        col_range.Delete()
                        print(f"    [{name}] 裁剪列: 保留 1-{max_c_keep}, 删除 {del_start_col}-{total_cols}")
                except Exception as e_col:
                    print(f"    [{name}] 列裁剪跳过: {e_col}")

                # 删除下方多余行
                try:
                    total_rows = new_ws.UsedRange.Rows.Count
                    if total_rows > max_r_keep:
                        del_start_row = max_r_keep + 1
                        row_range = new_ws.Range(
                            new_ws.Cells(del_start_row, 1),
                            new_ws.Cells(total_rows, 1)
                        ).EntireRow
                        row_range.Delete()
                        print(f"    [{name}] 裁剪行: 保留 1-{max_r_keep}, 删除 {del_start_row}-{total_rows}")
                except Exception as e_row:
                    print(f"    [{name}] 行裁剪跳过: {e_row}")

            html_path = str((tmp_dir / f"{safe_name}.htm").resolve())
            new_wb.SaveAs(html_path, FileFormat=44)  # 44 = HTML
            new_wb.Close(SaveChanges=False)

            # 打印导出文件大小
            htm_size = os.path.getsize(html_path)
            size_str = f"{htm_size / 1024:.0f}KB" if htm_size < 1024 * 1024 else f"{htm_size / 1024 / 1024:.1f}MB"
            result[name] = html_path
            print(f"  [{name}] HTML 导出完成 ✓ ({size_str})")
        except Exception as e:
            print(f"  [{name}] HTML 导出失败: {e}")

    return result


def _load_html_channel_data(html_path: str) -> Tuple[Dict[str, CellStyle], List[List[CellData]]]:
    """
    加载单个 sheet 的 HTML 通道数据：解析 CSS + 解析 HTML。
    返回 (css_classes, rows_data)。
    """
    htm_path = Path(html_path)
    files_dir = htm_path.parent / (htm_path.stem + ".files")

    # 查找 CSS 文件
    css_text = ""
    if files_dir.exists():
        for f in files_dir.iterdir():
            if f.suffix == '.css':
                try:
                    css_text = _read_html_auto_encoding(str(f))
                except Exception:
                    pass
                break

    # 查找实际的 sheet HTML 文件
    sheet_htm = None
    if files_dir.exists():
        for f in files_dir.iterdir():
            if f.name.startswith('sheet') and f.suffix in ('.htm', '.html'):
                sheet_htm = str(f)
                break

    if not sheet_htm:
        # 如果没有 .files 子目录，HTML 可能直接就是内容文件
        sheet_htm = html_path

    css_classes = _parse_css_classes(css_text) if css_text else {}
    rows_data = _parse_sheet_html(sheet_htm, css_classes)

    return css_classes, rows_data


# ═══════════════════════════════════════════════════════════
#  富文本格式化（用于 Markdown 输出）
# ═══════════════════════════════════════════════════════════

def _format_text_with_style(text: str, style: CellStyle, html_content: str = "") -> str:
    """
    将文本加上 HTML 格式标记。
    统一使用 HTML 标签而非 Markdown 语法，避免嵌套渲染失败。
    如果 html_content 包含内联混合格式，优先使用。
    """
    if not text.strip():
        return ""

    # 先检查是否有内联混合格式（同一单元格内部分文字样式不同）
    if html_content and ('<s>' in html_content or '<del>' in html_content
                          or '<b>' in html_content or '<i>' in html_content
                          or '<font' in html_content or '<u>' in html_content
                          or '<strong' in html_content or '<em' in html_content):
        formatted = _format_inline_html(html_content, style)
        if formatted:
            return formatted

    # 整个单元格统一格式
    result = text.strip()

    # 由内到外包裹：删除线 > 加粗 > 斜体 > 下划线 > 颜色 > 背景色
    if style.strikethrough:
        result = f'<del>{result}</del>'
    if style.bold:
        result = f'<strong>{result}</strong>'
    if style.italic:
        result = f'<em>{result}</em>'
    if style.underline:
        result = f'<u>{result}</u>'
    if style.color:
        result = f'<font color="{style.color}">{result}</font>'
    if style.bg_color:
        result = f'<span style="background-color:{style.bg_color}">{result}</span>'

    return result


def _format_inline_html(html_content: str, cell_style: CellStyle) -> str:
    """
    处理有内联格式标签的 HTML 内容（如单元格内部分文字有删除线/加粗/颜色混合）。

    支持嵌套标签递归处理：<font color="red"><b>文字</b></font> 会同时保留颜色和加粗。
    通过 inherited_styles 字典在递归中逐层累积样式。
    """
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
    except Exception:
        return ""

    result_parts = []

    def _wrap_with_styles(text: str, styles: dict) -> str:
        """按 styles 字典由内到外包裹格式标签"""
        result = text
        if styles.get('strikethrough'):
            result = f'<del>{result}</del>'
        if styles.get('bold'):
            result = f'<strong>{result}</strong>'
        if styles.get('italic'):
            result = f'<em>{result}</em>'
        if styles.get('underline'):
            result = f'<u>{result}</u>'
        color = styles.get('color')
        if color:
            result = f'<font color="{color}">{result}</font>'
        return result

    def process_node(node, inherited_styles: dict):
        if isinstance(node, str):
            text = node.replace('\u3000', '').strip()
            if text:
                result_parts.append(_wrap_with_styles(text, inherited_styles))
            return

        if not hasattr(node, 'name') or node.name is None:
            return

        # 拷贝继承样式，叠加当前节点的样式
        styles = dict(inherited_styles)

        if node.name in ('s', 'del', 'strike'):
            styles['strikethrough'] = True
        elif node.name in ('b', 'strong'):
            styles['bold'] = True
        elif node.name in ('i', 'em'):
            styles['italic'] = True
        elif node.name == 'u':
            styles['underline'] = True
        elif node.name == 'font':
            color = node.get('color', '')
            if color:
                nc = _normalize_color(color)
                if nc:
                    styles['color'] = nc
        elif node.name == 'br':
            result_parts.append('\n')
            return

        # 递归处理所有子节点
        if hasattr(node, 'children'):
            for child in node.children:
                process_node(child, styles)

    # 初始样式来自 cell_style（CSS class 定义的整体样式）
    initial_styles = {}
    if cell_style.color:
        initial_styles['color'] = cell_style.color

    for child in soup.children:
        process_node(child, initial_styles)

    return ''.join(result_parts)

# ── 自动裁剪工具 ──
def _auto_crop_whitespace(image_path: str, padding: int = 4, edge_threshold: int = 15) -> bool:
    """基于方向分离梯度裁剪截图空白区域。

    Excel COM 截图中行边框线会在所有列产生垂直梯度，干扰左右边界检测。
    因此：
    - 水平梯度 → 确定左右裁剪边界（不受行边框影响）
    - 垂直梯度 → 确定上下裁剪边界（不受列边框影响）
    """
    try:
        from PIL import ImageChops as _IC
        img = Image.open(image_path).convert('RGB')
        w, h = img.size
        gray = img.convert('L')
        hg = _IC.difference(gray.crop((0, 0, w - 1, h)), gray.crop((1, 0, w, h)))
        vg = _IC.difference(gray.crop((0, 0, w, h - 1)), gray.crop((0, 1, w, h)))
        def _pad_w(im):
            c = Image.new('L', (w, h), 0); c.paste(im, (0, 0)); return c
        def _pad_h(im):
            c = Image.new('L', (w, h), 0); c.paste(im, (0, 0)); return c
        hg_full = _pad_w(hg)
        vg_full = _pad_h(vg)

        # 水平梯度二值化 → 确定左右边界（不受行边框干扰）
        hg_bin = hg_full.point(lambda p: 255 if p > edge_threshold else 0)
        hg_bbox = hg_bin.getbbox()
        # 垂直梯度二值化 → 确定上下边界（不受列边框干扰）
        vg_bin = vg_full.point(lambda p: 255 if p > edge_threshold else 0)
        vg_bbox = vg_bin.getbbox()

        if not hg_bbox and not vg_bbox:
            return False

        # 从各方向梯度取对应维度的边界
        x1 = hg_bbox[0] if hg_bbox else 0
        x2 = hg_bbox[2] if hg_bbox else w
        y1 = vg_bbox[1] if vg_bbox else 0
        y2 = vg_bbox[3] if vg_bbox else h

        # 加 padding
        x1 = max(0, x1 - padding)
        y1 = max(0, y1 - padding)
        x2 = min(w, x2 + padding)
        y2 = min(h, y2 + padding)
        if (x2 - x1) * (y2 - y1) >= w * h * 0.95:
            return False
        img.crop((x1, y1, x2, y2)).save(image_path, 'PNG')
        return True
    except Exception:
        return False

# ──────────────────────────────────────────────
# 常量
# ──────────────────────────────────────────────
MAX_SCREENSHOT_WIDTH = 6000    # PNG 最大宽度像素
MAX_SCREENSHOT_HEIGHT = 6000   # PNG 最大高度像素
# 基于像素上限反算磅值阈值: 96 DPI 下 1pt ≈ 1.333px，所以 px / 1.333 ≈ pt
SPLIT_THRESHOLD_PTS = int(MAX_SCREENSHOT_HEIGHT / 1.333)  # 约 5250 pt
SCREENSHOT_PADDING_ROWS = 2   # 截图范围在内容边界外扩的安全行数
SCREENSHOT_PADDING_COLS = 3   # 截图范围在内容边界外扩的安全列数


# ──────────────────────────────────────────────
# 第一通道: 截图 (Excel COM)
# ──────────────────────────────────────────────

# 文字溢出估算的安全系数
# Excel 实际渲染受字体、字号、字间距影响，纯字符数估算偏保守
# 1.5x 系数可补偿大部分误差
TEXT_OVERFLOW_SAFETY_FACTOR = 1.8


def _estimate_text_visual_width(text: str) -> float:
    """
    估算文本的视觉宽度（以 Excel 默认字符单位计）。
    中文/全角字符按 2 个字符宽度计，英文/数字按 1 计。
    CJK 标点等全角符号也按 2 计。
    只取最长的一行（文本可能含换行符）。
    最终乘以安全系数以补偿字体渲染误差。
    """
    max_width = 0
    for line in str(text).split('\n'):
        w = 0
        for ch in line:
            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                w += 2  # CJK / 全角字符
            elif '\u2000' <= ch <= '\u206f':
                w += 1  # 通用标点
            else:
                w += 1
        max_width = max(max_width, w)
    return max_width * TEXT_OVERFLOW_SAFETY_FACTOR


def _get_col_width(ws_openpyxl, col: int) -> float:
    """获取列宽（字符单位），未设置则返回 Excel 默认值 8.43"""
    letter = get_column_letter(col)
    dim = ws_openpyxl.column_dimensions.get(letter)
    if dim and dim.width and dim.width > 0:
        return dim.width
    return 8.43  # Excel 默认列宽


def _calc_content_bounds(ws_openpyxl, images_info: List[dict]) -> Optional[Tuple[int, int, int, int]]:
    """
    用 openpyxl 精确计算工作表的真实内容边界（含嵌入图片区域 + 文字视觉溢出）。
    返回 (min_row, max_row, min_col, max_col)，无内容则返回 None。

    文字溢出处理：
    当单元格文字很长、未启用自动换行(wrap_text)、且右侧无合并单元格时，
    Excel 会让文字视觉上溢出到右边的空列。需要将这些溢出列也纳入截图范围。
    """
    min_r, max_r = float('inf'), 0
    min_c, max_c = float('inf'), 0

    # 预建合并单元格 lookup: (row, col) -> max_col_of_merge
    merge_spans = {}
    for mr in ws_openpyxl.merged_cells.ranges:
        merge_spans[(mr.min_row, mr.min_col)] = mr.max_col

    # 记录需要进行溢出估算的单元格: [(col, text_visual_width, cell_col_width)]
    overflow_candidates = []

    # 1. 扫描所有有内容的单元格
    if ws_openpyxl.max_row and ws_openpyxl.max_column:
        for row in ws_openpyxl.iter_rows(
            min_row=1, max_row=min(ws_openpyxl.max_row, 2000),
            min_col=1, max_col=min(ws_openpyxl.max_column, 200)
        ):
            for cell in row:
                if cell.value is not None:
                    min_r = min(min_r, cell.row)
                    max_r = max(max_r, cell.row)
                    min_c = min(min_c, cell.column)
                    max_c = max(max_c, cell.column)

                    # 检查是否需要估算文字溢出
                    text = str(cell.value)
                    if len(text) <= 6:
                        continue  # 短文本不会溢出

                    # 如果启用了自动换行，文字不会溢出
                    is_wrapped = (cell.alignment and cell.alignment.wrap_text)
                    if is_wrapped:
                        continue

                    # 如果是合并单元格的左上角，用合并后的总宽度
                    merge_end_col = merge_spans.get((cell.row, cell.column))
                    if merge_end_col:
                        # 计算合并区域总宽度
                        total_col_w = sum(
                            _get_col_width(ws_openpyxl, c)
                            for c in range(cell.column, merge_end_col + 1)
                        )
                        effective_start_col = merge_end_col
                    else:
                        total_col_w = _get_col_width(ws_openpyxl, cell.column)
                        effective_start_col = cell.column

                    text_w = _estimate_text_visual_width(text)
                    if text_w > total_col_w:
                        overflow_candidates.append((effective_start_col, text_w - total_col_w))

    # 2. 计算文字溢出导致的最大右边界
    if overflow_candidates:
        max_visual_col = max_c
        for start_col, overflow_chars in overflow_candidates:
            # 从溢出起始列向右累积列宽，直到容纳溢出字符
            remaining = overflow_chars
            col_cursor = start_col + 1
            while remaining > 0 and col_cursor <= 500:  # 安全上限
                col_w = _get_col_width(ws_openpyxl, col_cursor)
                remaining -= col_w
                col_cursor += 1
            visual_end_col = col_cursor - 1
            max_visual_col = max(max_visual_col, visual_end_col)

        # 限制溢出扩展：最多比原始 max_c 多 30 列，避免极端情况
        max_overflow_limit = max_c + 30
        if max_visual_col > max_c:
            old_max_c = max_c
            max_c = min(max_visual_col, max_overflow_limit)
            print(f"    [边界] 文字溢出扩展: 列{old_max_c} → 列{max_c} (+{max_c - old_max_c}列)")

    # 3. 考虑嵌入图片的位置（行和列）
    for img in images_info:
        sr = img.get('start_row')
        er = img.get('end_row')
        sc = img.get('start_col')
        ec = img.get('end_col')
        if sr is not None:
            min_r = min(min_r, sr)
            max_r = max(max_r, sr)
        if er is not None:
            max_r = max(max_r, er)
        if sc is not None and min_c != float('inf'):
            min_c = min(min_c, sc)
        if ec is not None and max_c > 0:
            max_c = max(max_c, ec)

    if max_r == 0:
        return None
    if min_c == float('inf'):
        min_c = 1
    if max_c == 0:
        max_c = 1

    return (int(min_r), int(max_r), int(min_c), int(max_c))


# 行范围截图时的列方向安全距离（单元格数）
ROW_RANGE_COL_SAFETY = 1


def _calc_row_range_col_bounds(ws_openpyxl, start_row: int, end_row: int,
                               images_info: List[dict],
                               max_col_override: Optional[int] = None) -> Optional[Tuple[int, int]]:
    """
    计算指定行范围内的实际内容列边界。
    不做文字溢出估算——列范围由 Agent 分析截图后决定。

    Args:
        max_col_override: Agent 指定的最大列号。如果提供，直接使用（+ 安全距离），
                         忽略 openpyxl 的计算结果。

    返回 (min_col, max_col)，无内容则返回 None。
    """
    min_c, max_c = float('inf'), 0

    # 扫描行范围内有值的单元格，确定实际内容列边界
    for row in ws_openpyxl.iter_rows(
        min_row=start_row, max_row=min(end_row, ws_openpyxl.max_row or end_row),
        min_col=1, max_col=min(ws_openpyxl.max_column or 200, 200)
    ):
        for cell in row:
            if cell.value is not None:
                min_c = min(min_c, cell.column)
                max_c = max(max_c, cell.column)

    # 考虑合并单元格的右边界
    for mr in ws_openpyxl.merged_cells.ranges:
        if mr.min_row <= end_row and mr.max_row >= start_row:
            if mr.min_col >= min_c:
                max_c = max(max_c, mr.max_col)

    if max_c == 0:
        return None

    # 如果 Agent 指定了列范围，优先使用
    if max_col_override is not None:
        max_c = max(max_c, max_col_override)

    # 嵌入图片的列范围（图片与行范围有交集即纳入）
    for img in images_info:
        sr = img.get('start_row')
        er = img.get('end_row') or sr
        if sr is not None and er is not None:
            # 图片行范围 [sr, er] 与目标行范围 [start_row, end_row] 有交集
            if sr <= end_row and er >= start_row:
                ec = img.get('end_col')
                sc = img.get('start_col')
                if ec is not None:
                    max_c = max(max_c, ec)
                if sc is not None:
                    min_c = min(min_c, sc)

    # 加上单元格级安全距离
    max_c += ROW_RANGE_COL_SAFETY

    return (int(min_c), int(max_c))


def _export_sheet_screenshot(ws_com, output_dir: Path, sheet_name: str,
                             content_bounds: Optional[Tuple[int, int, int, int]] = None,
                             row_range: Optional[Tuple[int, int]] = None,
                             output_filename: Optional[str] = None) -> List[str]:
    """
    将工作表导出为 PNG 截图。
    如果提供了 content_bounds (min_row, max_row, min_col, max_col)，
    则只截取内容区域 + 安全边距，避免大片空白。
    对于特别大的 Sheet，按行分片导出。

    Args:
        row_range: 可选 (start_row, end_row)，指定截图的行范围。
                   当指定时，覆盖 content_bounds 的行范围，列范围仍从 content_bounds 取。
                   超出实际内容边界时自动裁剪。
        output_filename: 可选，自定义输出文件名（如 "module_1.png"）。
                         未指定时使用默认命名 "screenshot.png" 或 "screenshot_partN.png"。

    返回生成的文件名列表。
    """
    screenshots = []

    try:
        used = ws_com.UsedRange
        if used is None or used.Rows.Count == 0:
            print(f"    [截图] {sheet_name}: 无内容，跳过")
            return screenshots

        # ── 确定截图范围 ──
        if content_bounds:
            cb_min_r, cb_max_r, cb_min_c, cb_max_c = content_bounds

            # 如果指定了 row_range，用其覆盖行范围
            if row_range:
                rr_start, rr_end = row_range
                # 裁剪到 content_bounds 边界
                cb_min_r = max(cb_min_r, rr_start)
                cb_max_r = min(cb_max_r, rr_end)
                if cb_min_r > cb_max_r:
                    print(f"    [截图] {sheet_name}: row_range ({rr_start}-{rr_end}) 超出内容边界，跳过")
                    return screenshots

            # 外扩安全边距（指定 row_range 时行方向不加 padding，避免相邻模块截图重叠）
            if row_range:
                snap_min_r = cb_min_r
                snap_max_r = cb_max_r
            else:
                snap_min_r = max(1, cb_min_r - SCREENSHOT_PADDING_ROWS)
                snap_max_r = cb_max_r + SCREENSHOT_PADDING_ROWS
            snap_min_c = max(1, cb_min_c - SCREENSHOT_PADDING_COLS)
            snap_max_c = cb_max_c + SCREENSHOT_PADDING_COLS

            # 行方向不超过 UsedRange（避免截到空行）
            # 列方向不限制：文字溢出可能超出 UsedRange 的列范围
            used_max_r = used.Row + used.Rows.Count - 1
            snap_max_r = min(snap_max_r, used_max_r)

            target_range = ws_com.Range(
                ws_com.Cells(snap_min_r, snap_min_c),
                ws_com.Cells(snap_max_r, snap_max_c)
            )

            orig_w = used.Width
            orig_h = used.Height
            new_w = target_range.Width
            new_h = target_range.Height
            saved_pct = (1 - (new_w * new_h) / (orig_w * orig_h)) * 100 if orig_w * orig_h > 0 else 0

            print(f"    [截图] {sheet_name}: 内容区域 R{cb_min_r}:R{cb_max_r} C{cb_min_c}:C{cb_max_c}"
                  f" → 截图 R{snap_min_r}:R{snap_max_r} C{snap_min_c}:C{snap_max_c}"
                  f" (节省 {saved_pct:.0f}% 面积)")
        else:
            # 无 bounds 信息，回退到 UsedRange
            target_range = used
            snap_min_r = used.Row
            snap_max_r = used.Row + used.Rows.Count - 1
            snap_min_c = used.Column
            snap_max_c = used.Column + used.Columns.Count - 1
            print(f"    [截图] {sheet_name}: 使用 UsedRange {used.Width:.0f}x{used.Height:.0f} pt")

        content_height = target_range.Height
        content_width = target_range.Width
        print(f"    [截图] {sheet_name}: {content_width:.0f} x {content_height:.0f} pt, "
              f"{snap_max_c - snap_min_c + 1} 列 x {snap_max_r - snap_min_r + 1} 行")

        # 确定输出文件名
        default_single_name = output_filename if output_filename else "screenshot.png"

        if content_height <= SPLIT_THRESHOLD_PTS:
            # 整个范围一张截图
            fname = _copy_range_as_image(ws_com, target_range, output_dir, default_single_name)
            if fname:
                screenshots.append(fname)
        else:
            # 分片截图
            start_row = snap_min_r
            end_row = snap_max_r
            part = 1

            current_row = start_row
            while current_row <= end_row:
                # 累积行直到高度达到阈值
                chunk_end = current_row
                accumulated_h = 0
                while chunk_end <= end_row:
                    row_h = ws_com.Rows(chunk_end).Height
                    if accumulated_h + row_h > SPLIT_THRESHOLD_PTS and chunk_end > current_row:
                        break
                    accumulated_h += row_h
                    chunk_end += 1
                chunk_end -= 1  # 回退到最后一个包含的行

                rng = ws_com.Range(
                    ws_com.Cells(current_row, snap_min_c),
                    ws_com.Cells(chunk_end, snap_max_c)
                )
                part_name = output_filename.replace('.png', f'_part{part}.png') if output_filename else f"screenshot_part{part}.png"
                fname = _copy_range_as_image(
                    ws_com, rng, output_dir, part_name
                )
                if fname:
                    screenshots.append(fname)
                part += 1
                current_row = chunk_end + 1

    except Exception as e:
        print(f"    [截图] {sheet_name} 失败: {e}")

    return screenshots


def _crop_page_whitespace(img, crop_top=True, crop_bottom=True):
    """
    裁剪单页 PDF 渲染图的上下空白区域（保留左右完整宽度）。
    用于多页拼接前去除页面间的多余白边。

    Args:
        img: PIL Image 对象
        crop_top: 是否裁剪顶部空白（首页可设为 False 保留顶部）
        crop_bottom: 是否裁剪底部空白（末页可设为 False 保留底部）

    Returns:
        裁剪后的 PIL Image 对象（如果无法检测到内容则返回原图）
    """
    try:
        gray = img.convert('L')
        # 将非白色像素标记为内容（阈值 248，略低于纯白以捕获浅灰边缘）
        bw = gray.point(lambda p: 0 if p >= 248 else 255)
        # 按行求和，找到有内容的行范围
        w = bw.width
        pixels = list(bw.getdata())
        row_sums = [sum(pixels[y * w:(y + 1) * w]) for y in range(bw.height)]

        # 找第一行和最后一行有内容的位置
        first_row = 0
        last_row = len(row_sums) - 1
        for i, s in enumerate(row_sums):
            if s > 0:
                first_row = i
                break
        for i in range(len(row_sums) - 1, -1, -1):
            if row_sums[i] > 0:
                last_row = i
                break

        # 根据裁剪模式决定边距
        if crop_top:
            y1 = first_row  # 紧贴内容，无额外边距
        else:
            y1 = 0  # 保留顶部原样

        if crop_bottom:
            y2 = last_row + 1  # 紧贴内容，无额外边距
        else:
            y2 = img.height  # 保留底部原样

        if y2 - y1 < img.height * 0.3:
            # 裁剪超过 70% 说明检测可能有误，返回原图
            return img

        return img.crop((0, y1, img.width, y2))
    except Exception:
        return img


def _copy_range_as_image_via_pdf(ws_com, rng, output_dir: Path, filename: str) -> Optional[str]:
    """
    利用 Excel COM 导出 PDF 再渲染为 PNG 的方式截图。

    流程：
    1. 设置 PrintArea 为指定 Range
    2. 配置 PageSetup：零边距、100% 原尺寸（不缩放）、横向布局
    3. ExportAsFixedFormat 导出临时 PDF（可能多页）
    4. PyMuPDF (fitz) 逐页高清渲染、裁剪页间空白、垂直拼接
    5. 清理临时 PDF

    关键设计：
    - FitToPages 1×1：所有内容强制缩放到单页，彻底消除 Excel 分页导致的间隙/行重复
    - A3 横向纸张：1190pt 可用宽度，减少缩放幅度
    - 动态渲染倍率：根据缩放比例自动计算 render scale，保证最终 ≥150 DPI
    - 自动分片：内容过高时（fit_scale < 阈值），按行拆分为多个子区域，
      每个子区域独立导出单页 PDF，最终 PIL 拼接。与 Excel 自动分页不同，
      自行分片通过设置不同 PrintArea 实现，不会产生行重复。
    """
    # A3 横向可用区域（零边距）
    PAGE_W, PAGE_H = 1190.0, 841.0
    # fit_scale 低于此值时触发自动分片（确保 render_scale ≤ 8x 时仍能达到 ~150 DPI）
    # 150 / (72 * 8) = 0.26，取 0.30 留余量
    MIN_FIT_SCALE = 0.30
    TARGET_DPI = 150

    out_file = str(output_dir / filename)
    pdf_file = str(output_dir / filename.replace('.png', '_temp.pdf'))

    try:
        ps = ws_com.PageSetup
        orig_print_area = ps.PrintArea
        orig_zoom = ps.Zoom
        orig_fit_wide = ps.FitToPagesWide
        orig_fit_tall = ps.FitToPagesTall
        orig_left = ps.LeftMargin
        orig_right = ps.RightMargin
        orig_top = ps.TopMargin
        orig_bottom = ps.BottomMargin
        orig_header = ps.HeaderMargin
        orig_footer = ps.FooterMargin
        orig_orientation = ps.Orientation
        orig_paper_size = ps.PaperSize
        orig_gridlines = ps.PrintGridlines
        orig_display_gridlines = ws_com.Parent.Windows(1).DisplayGridlines if ws_com.Parent.Windows.Count > 0 else True

        try:
            range_w = rng.Width
            range_h = rng.Height
            fit_scale_w = min(PAGE_W / max(range_w, 1), 1.0)
            fit_scale_full = min(fit_scale_w, PAGE_H / max(range_h, 1), 1.0)

            # 判断是否需要分片
            if fit_scale_full >= MIN_FIT_SCALE:
                # ── 单片模式：整个 Range 一次导出 ──
                chunks = [(rng, range_w, range_h)]
            else:
                # ── 分片模式：按行拆分，每片保证 fit_scale >= MIN_FIT_SCALE ──
                # 每片最大高度：PAGE_H / MIN_FIT_SCALE（但不超过 PAGE_H / fit_scale_w，因为宽度缩放也要考虑）
                max_chunk_h = PAGE_H / max(MIN_FIT_SCALE, fit_scale_w)
                # 获取 Range 的行列范围
                first_row = rng.Row
                last_row = first_row + rng.Rows.Count - 1
                first_col = rng.Column
                last_col = first_col + rng.Columns.Count - 1

                chunks = []
                cur_row = first_row
                while cur_row <= last_row:
                    # 累积行直到高度达到 max_chunk_h
                    chunk_end = cur_row
                    acc_h = 0.0
                    while chunk_end <= last_row:
                        row_h = ws_com.Rows(chunk_end).Height
                        if acc_h + row_h > max_chunk_h and chunk_end > cur_row:
                            break
                        acc_h += row_h
                        chunk_end += 1
                    chunk_end -= 1

                    chunk_rng = ws_com.Range(
                        ws_com.Cells(cur_row, first_col),
                        ws_com.Cells(chunk_end, last_col)
                    )
                    chunks.append((chunk_rng, chunk_rng.Width, chunk_rng.Height))
                    cur_row = chunk_end + 1

                print(f"    [截图-PDF] 内容过高 ({range_h:.0f}pt, fit_scale={fit_scale_full:.2f})，"
                      f"自动分为 {len(chunks)} 片")

            # ── 通用 PageSetup（所有片共享）──
            ps.LeftMargin = 0
            ps.RightMargin = 0
            ps.TopMargin = 0
            ps.BottomMargin = 0
            ps.HeaderMargin = 0
            ps.FooterMargin = 0
            ps.PrintGridlines = False  # 关闭打印网格线
            # 同时关闭显示网格线，某些 Excel 版本 PDF 导出会参考此设置
            try:
                if ws_com.Parent.Windows.Count > 0:
                    ws_com.Parent.Windows(1).DisplayGridlines = False
            except Exception:
                pass
            try:
                ps.PaperSize = 8  # xlPaperA3
            except Exception:
                pass
            ps.Orientation = 2  # xlLandscape
            ps.Zoom = False
            ps.FitToPagesWide = 1
            ps.FitToPagesTall = 1
            try:
                ps.LeftHeader = ""
                ps.CenterHeader = ""
                ps.RightHeader = ""
                ps.LeftFooter = ""
                ps.CenterFooter = ""
                ps.RightFooter = ""
            except Exception:
                pass

            import fitz  # PyMuPDF
            from PIL import Image as _PILImage

            chunk_images = []

            for ci, (chunk_rng, cw, ch) in enumerate(chunks):
                # 设置当前片的打印区域
                ps.PrintArea = chunk_rng.Address

                # 导出 PDF
                chunk_pdf = str(output_dir / filename.replace('.png', f'_chunk{ci}.pdf'))
                try:
                    ws_com.ExportAsFixedFormat(
                        Type=0, Filename=chunk_pdf, Quality=0,
                        IncludeDocProperties=False, IgnorePrintAreas=False,
                        OpenAfterPublish=False,
                    )
                    if not os.path.exists(chunk_pdf) or os.path.getsize(chunk_pdf) == 0:
                        print(f"    [截图-PDF] 片{ci} PDF 导出为空")
                        continue

                    doc = fitz.open(chunk_pdf)
                    if doc.page_count == 0:
                        doc.close()
                        continue

                    # 计算当前片的渲染倍率
                    chunk_fit = min(PAGE_W / max(cw, 1), PAGE_H / max(ch, 1), 1.0)
                    rs = TARGET_DPI / (72.0 * chunk_fit)
                    rs = max(rs, 2.0)
                    rs = min(rs, 8.0)
                    eff_dpi = 72.0 * rs * chunk_fit

                    # 关闭图形抗锯齿（消除单元格边界伪影线），保留文字抗锯齿
                    try:
                        fitz.mupdf.fz_set_graphics_aa_level(0)  # 关闭图形AA
                        fitz.mupdf.fz_set_text_aa_level(8)      # 保留文字AA
                    except Exception:
                        try:
                            fitz.TOOLS.set_aa_level(0)  # fallback: 全部关闭
                        except Exception:
                            pass
                    mat = fitz.Matrix(rs, rs)
                    page = doc[0]
                    pix = page.get_pixmap(matrix=mat, alpha=False)
                    try:
                        fitz.mupdf.fz_set_graphics_aa_level(8)  # 恢复
                        fitz.mupdf.fz_set_text_aa_level(8)
                    except Exception:
                        try:
                            fitz.TOOLS.set_aa_level(8)
                        except Exception:
                            pass
                    img = _PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    doc.close()

                    if len(chunks) > 1:
                        print(f"    [截图-PDF] 片{ci}: fit={chunk_fit:.2f}, "
                              f"render={rs:.1f}x, ≈{eff_dpi:.0f}DPI, "
                              f"{img.width}x{img.height}px")
                    else:
                        print(f"    [截图-PDF] 单页模式: fit={chunk_fit:.2f}, "
                              f"render={rs:.1f}x, ≈{eff_dpi:.0f}DPI")

                    chunk_images.append(img)
                finally:
                    try:
                        if os.path.exists(chunk_pdf):
                            os.remove(chunk_pdf)
                    except Exception:
                        pass

            if not chunk_images:
                print(f"    [截图-PDF] 所有片段导出失败: {filename}")
                return None

            if len(chunk_images) == 1:
                chunk_images[0].save(out_file, 'PNG')
            else:
                # 多片拼接：每片独立 PrintArea 导出，无行重复，直接裁剪空白后拼接
                cropped = []
                for i, img in enumerate(chunk_images):
                    img = _crop_page_whitespace(img, crop_top=True, crop_bottom=True)
                    cropped.append(img)

                total_w = max(im.width for im in cropped)
                total_h = sum(im.height for im in cropped)
                merged = _PILImage.new('RGB', (total_w, total_h), (255, 255, 255))
                y = 0
                for im in cropped:
                    merged.paste(im, (0, y))
                    y += im.height
                merged.save(out_file, 'PNG')

            if os.path.exists(out_file) and os.path.getsize(out_file) > 0:
                _auto_crop_whitespace(out_file)
                size_kb = os.path.getsize(out_file) / 1024
                print(f"    [截图-PDF] 已保存: {filename} ({size_kb:.1f} KB)")
                return filename
            else:
                print(f"    [截图-PDF] PNG 渲染结果为空: {filename}")
                return None

        finally:
            try:
                ps.PrintArea = orig_print_area if orig_print_area else ""
                ps.FitToPagesWide = orig_fit_wide if orig_fit_wide else False
                ps.FitToPagesTall = orig_fit_tall if orig_fit_tall else False
                ps.Zoom = orig_zoom if orig_zoom else 100
                ps.LeftMargin = orig_left
                ps.RightMargin = orig_right
                ps.TopMargin = orig_top
                ps.BottomMargin = orig_bottom
                ps.HeaderMargin = orig_header
                ps.FooterMargin = orig_footer
                ps.Orientation = orig_orientation
                ps.PaperSize = orig_paper_size
                ps.PrintGridlines = orig_gridlines
                try:
                    if ws_com.Parent.Windows.Count > 0:
                        ws_com.Parent.Windows(1).DisplayGridlines = orig_display_gridlines
                except Exception:
                    pass
            except Exception as e:
                print(f"    [截图-PDF] 恢复 PageSetup 失败（不影响截图结果）: {e}")

            # 清理可能残留的临时文件
            for f in output_dir.glob(filename.replace('.png', '_chunk*.pdf')):
                try:
                    f.unlink()
                except Exception:
                    pass
            try:
                if os.path.exists(pdf_file):
                    os.remove(pdf_file)
            except Exception:
                pass

    except Exception as e:
        print(f"    [截图-PDF] 导出失败 {filename}: {e}")
        for f in output_dir.glob(filename.replace('.png', '_chunk*.pdf')):
            try:
                f.unlink()
            except Exception:
                pass
        return None


def _copy_range_as_image_clipboard(ws_com, rng, output_dir: Path, filename: str) -> Optional[str]:
    """
    利用 Excel COM 将指定 Range 复制为图片，通过剪贴板直接保存为 PNG。
    （CopyPicture 方式，作为 PDF 方式的回退方案）

    流程：CopyPicture → 剪贴板 → PIL.ImageGrab.grabclipboard() → 保存 PNG
    包含 COM retry + message pumping 以应对异步剪贴板操作。
    """
    MAX_RETRIES = 3
    out_file = str(output_dir / filename)

    for attempt in range(MAX_RETRIES):
        try:
            # 清空剪贴板
            import win32clipboard
            try:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
            except Exception:
                pass

            # 确保目标 sheet 已激活（解决 CopyPicture 操作到非激活页导致失败的问题）
            try:
                ws_com.Activate()
            except Exception:
                pass
            import time
            time.sleep(1)

            # CopyPicture: Appearance=xlScreen(1), Format=xlBitmap(2)
            rng.CopyPicture(Appearance=1, Format=2)

            # Message pumping：让 COM 完成异步剪贴板操作
            time.sleep(0.5)
            pythoncom.PumpWaitingMessages()

            # 从剪贴板抓图
            from PIL import ImageGrab
            cb_img = None
            for wait_attempt in range(5):
                cb_img = ImageGrab.grabclipboard()
                if cb_img is not None:
                    break
                time.sleep(0.2)
                pythoncom.PumpWaitingMessages()

            if cb_img is None:
                if attempt < MAX_RETRIES - 1:
                    print(f"    [截图] 剪贴板为空，重试 ({attempt + 1}/{MAX_RETRIES})...")
                    time.sleep(0.5)
                    continue
                print(f"    [截图] 剪贴板抓取失败: {filename}")
                return None

            # 保存 PNG
            if cb_img.mode == 'RGBA':
                # 去除 alpha 通道，白底合成
                bg = Image.new('RGB', cb_img.size, (255, 255, 255))
                bg.paste(cb_img, mask=cb_img.split()[3])
                cb_img = bg
            elif cb_img.mode != 'RGB':
                cb_img = cb_img.convert('RGB')

            cb_img.save(out_file, 'PNG')

            # 检查文件是否生成
            if os.path.exists(out_file) and os.path.getsize(out_file) > 0:
                _auto_crop_whitespace(out_file)
                size_kb = os.path.getsize(out_file) / 1024
                print(f"    [截图] 已保存: {filename} ({size_kb:.1f} KB)")
                return filename
            else:
                if attempt < MAX_RETRIES - 1:
                    print(f"    [截图] 文件为空，重试 ({attempt + 1}/{MAX_RETRIES})...")
                    continue
                print(f"    [截图] 生成失败: {filename}")
                return None

        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print(f"    [截图] 异常 ({attempt + 1}/{MAX_RETRIES}): {e}")
                import time
                time.sleep(0.5)
                continue
            print(f"    [截图] 导出失败 {filename}: {e}")
            return None

    return None


def _copy_range_as_image(ws_com, rng, output_dir: Path, filename: str) -> Optional[str]:
    """
    将指定 Range 导出为 PNG 截图。

    优先使用 PDF 导出方式（更稳定），失败时回退到 CopyPicture 剪贴板方式。
    """
    # 优先尝试 PDF 方式
    result = _copy_range_as_image_via_pdf(ws_com, rng, output_dir, filename)
    if result:
        return result

    # PDF 方式失败，回退到 CopyPicture 剪贴板方式
    print(f"    [截图] PDF 方式失败，回退到 CopyPicture 方式: {filename}")
    return _copy_range_as_image_clipboard(ws_com, rng, output_dir, filename)


# ──────────────────────────────────────────────
# 第二通道: 结构化文本 (openpyxl)
# ──────────────────────────────────────────────

def _color_to_hex(color) -> Optional[str]:
    """将 openpyxl 颜色对象转为 #RRGGBB 字符串"""
    if color is None:
        return None
    if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
        rgb = str(color.rgb)
        if len(rgb) == 8:  # AARRGGBB
            return f"#{rgb[2:]}"
        elif len(rgb) == 6:
            return f"#{rgb}"
    if hasattr(color, 'theme') and color.theme is not None:
        # 主题色，无法精确还原，标注主题编号
        return f"theme({color.theme})"
    return None


def _describe_font(font) -> List[str]:
    """描述字体的特殊格式"""
    attrs = []
    if font.bold:
        attrs.append("加粗")
    if font.italic:
        attrs.append("斜体")
    if font.strikethrough:
        attrs.append("删除线")
    if font.underline and font.underline != 'none':
        attrs.append("下划线")
    color_hex = _color_to_hex(font.color)
    if color_hex and color_hex not in ('#000000', 'theme(1)'):
        attrs.append(f"字色={color_hex}")
    if font.size and font.size != 11:
        attrs.append(f"字号={font.size}")
    return attrs


def _describe_fill(fill) -> Optional[str]:
    """描述单元格填充/背景色"""
    if fill.fill_type and fill.fill_type != 'none':
        fg = _color_to_hex(fill.fgColor)
        if fg and fg not in ('#000000',):
            return fg
    return None


def _get_merged_ranges(ws_openpyxl) -> dict:
    """
    返回 {(row, col): "合并区域描述"} 的字典，
    仅记录合并区域的左上角单元格。
    """
    merged = {}
    for merge_range in ws_openpyxl.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        max_row, max_col = merge_range.max_row, merge_range.max_col
        rows_span = max_row - min_row + 1
        cols_span = max_col - min_col + 1
        if rows_span > 1 or cols_span > 1:
            merged[(min_row, min_col)] = f"合并{rows_span}行x{cols_span}列"
    return merged


def _build_image_row_index(images_info: List[dict]) -> dict:
    """
    将嵌入图片按行号区间索引。
    返回 {row_number: [image_info_dict, ...]} —— 仅索引每张图片起始行。
    """
    row_index = {}
    for img in images_info:
        start_row = img.get('start_row')
        if start_row is not None:
            row_index.setdefault(start_row, []).append(img)
    return row_index


# ═══════════════════════════════════════════════════════════
#  表格区域检测 + 置信度分级
# ═══════════════════════════════════════════════════════════

# 置信度级别
CONFIDENCE_HIGH = "high"       # 直接生成 Markdown 表格
CONFIDENCE_MEDIUM = "medium"   # 生成初版表格 + 标记请 Agent 复核
CONFIDENCE_SUSPECT = "suspect" # 不自动转换，标记区域请 Agent 判断


@dataclass
class TableRegion:
    """检测到的表格区域"""
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    confidence: str = CONFIDENCE_HIGH  # high / medium / suspect
    reason: str = ""                    # 置信度判定原因
    has_border: bool = False            # HTML 通道检测到边框


def _detect_border_tables_openpyxl(ws_openpyxl, max_row: int, max_col: int) -> List[TableRegion]:
    """
    基于 openpyxl 原生边框信息检测表格区域（不依赖 HTML 通道）。
    openpyxl 直接读取 cell.border 的 top/bottom/left/right style，
    检测到的边框行集合与 HTML 通道完全一致，且单元格数量更多。
    """
    border_rows = set()
    border_cells = {}  # {row: set(col)}

    for row in ws_openpyxl.iter_rows(min_row=1, max_row=max_row,
                                      min_col=1, max_col=max_col):
        for cell in row:
            try:
                b = cell.border
                if b and any(
                    getattr(b, side) and getattr(b, side).style
                    for side in ('top', 'bottom', 'left', 'right')
                ):
                    border_rows.add(cell.row)
                    if cell.row not in border_cells:
                        border_cells[cell.row] = set()
                    border_cells[cell.row].add(cell.column)
            except Exception:
                pass

    if not border_rows:
        return []

    # 合并相邻边框行（允许最多 2 行 gap）
    sorted_rows = sorted(border_rows)
    raw_regions = []
    region_start = sorted_rows[0]
    region_end = sorted_rows[0]
    gap_count = 0

    for r in sorted_rows[1:]:
        gap = r - region_end - 1
        if gap <= 2:
            if gap > 0:
                gap_count += 1
            region_end = r
        else:
            raw_regions.append((region_start, region_end, gap_count))
            region_start = r
            region_end = r
            gap_count = 0

    raw_regions.append((region_start, region_end, gap_count))

    # 对每个行区域，按列间隙拆分成独立的列簇，每个簇生成独立 TableRegion
    results = []
    for (rs, re_, gaps) in raw_regions:
        row_count = re_ - rs + 1
        if row_count < 2:
            continue

        # 收集该行区域内所有有边框的列号
        all_border_cols = set()
        for r in range(rs, re_ + 1):
            if r in border_cells:
                all_border_cols.update(border_cells[r])

        if not all_border_cols:
            continue

        # ── 按列间隙拆分列簇（gap >= 2 列无边框则拆分）──
        col_clusters = _split_column_clusters(sorted(all_border_cols), gap_threshold=2)

        for (cl_min, cl_max) in col_clusters:
            # 进一步验证：该列簇在多少行有边框（排除偶尔飞入的孤立列）
            rows_with_border_in_cluster = 0
            for r in range(rs, re_ + 1):
                if r in border_cells and border_cells[r] & set(range(cl_min, cl_max + 1)):
                    rows_with_border_in_cluster += 1
            if rows_with_border_in_cluster < 2:
                continue

            # ── 启发式：区分数据表格 vs UI 线框图 ──
            region_info = _evaluate_table_region(
                ws_openpyxl, rs, re_, cl_min, cl_max, row_count, gaps
            )
            results.append(region_info)

    return results


def _split_column_clusters(sorted_cols: List[int], gap_threshold: int = 2) -> List[tuple]:
    """
    将有序列号列表按间隙拆分成独立的列簇。
    gap_threshold: 连续无边框列数 >= 此值则拆分。
    返回 [(col_min, col_max), ...] 列表。
    """
    if not sorted_cols:
        return []

    clusters = []
    cluster_start = sorted_cols[0]
    cluster_end = sorted_cols[0]

    for c in sorted_cols[1:]:
        if c - cluster_end <= gap_threshold:
            cluster_end = c
        else:
            clusters.append((cluster_start, cluster_end))
            cluster_start = c
            cluster_end = c

    clusters.append((cluster_start, cluster_end))
    return clusters


def _evaluate_table_region(ws_openpyxl, rs: int, re_: int,
                            col_min: int, col_max: int,
                            row_count: int, gaps: int) -> 'TableRegion':
    """
    对单个候选表格区域进行启发式评估（线框图检测 + 置信度分级）。
    """
    wireframe_signals = []
    col_span = col_max - col_min + 1
    total_cells = row_count * col_span

    # 信号1: 填充率 — 统计区域内有内容的单元格比例
    filled_count = 0
    for r in range(rs, re_ + 1):
        for c in range(col_min, col_max + 1):
            cell = ws_openpyxl.cell(row=r, column=c)
            if cell.value is not None and str(cell.value).strip():
                filled_count += 1
    fill_rate = filled_count / total_cells if total_cells > 0 else 0

    if fill_rate < 0.25:
        wireframe_signals.append(f"低填充率{fill_rate:.0%}")

    # 信号2: 列一致性 — 数据表格每行填充相同的列位置，线框图则不规则
    row_col_patterns = []
    for r in range(rs, re_ + 1):
        cols_with_content = frozenset(
            c for c in range(col_min, col_max + 1)
            if ws_openpyxl.cell(row=r, column=c).value is not None
            and str(ws_openpyxl.cell(row=r, column=c).value).strip()
        )
        if cols_with_content:
            row_col_patterns.append(cols_with_content)

    if len(row_col_patterns) >= 3:
        from collections import Counter
        pattern_counter = Counter(row_col_patterns)
        most_common_count = pattern_counter.most_common(1)[0][1]
        pattern_consistency = most_common_count / len(row_col_patterns)
        if pattern_consistency < 0.3:
            wireframe_signals.append(f"列模式不一致{pattern_consistency:.0%}")

    # 信号3: 合并单元格密度 — 线框图常有大量不规则合并
    merge_count = 0
    large_merge_count = 0
    for mg in ws_openpyxl.merged_cells.ranges:
        mr_min, mr_max = mg.min_row, mg.max_row
        mc_min, mc_max = mg.min_col, mg.max_col
        if mr_min >= rs and mr_max <= re_ and mc_min >= col_min and mc_max <= col_max:
            merge_count += 1
            merge_span = (mr_max - mr_min + 1) * (mc_max - mc_min + 1)
            if merge_span >= 4:
                large_merge_count += 1

    merge_density = merge_count / row_count if row_count > 0 else 0
    if merge_density > 0.5 or large_merge_count >= 3:
        wireframe_signals.append(f"高合并密度({merge_count}个合并,{large_merge_count}个大合并)")

    # 综合判定
    if len(wireframe_signals) >= 2:
        confidence = CONFIDENCE_SUSPECT
        reason = f"疑似UI线框图({row_count}行, {'; '.join(wireframe_signals)})"
    elif len(wireframe_signals) == 1:
        confidence = CONFIDENCE_MEDIUM
        reason = f"边框区域可能为线框图({row_count}行, {wireframe_signals[0]})"
    elif row_count >= 3 and gaps == 0:
        confidence = CONFIDENCE_HIGH
        reason = f"完整边框表格({row_count}行, 无gap)"
    elif row_count >= 3:
        confidence = CONFIDENCE_MEDIUM
        reason = f"边框表格({row_count}行, {gaps}个gap)"
    else:
        confidence = CONFIDENCE_MEDIUM
        reason = f"小边框表格({row_count}行)"

    return TableRegion(
        start_row=rs,
        end_row=re_,
        start_col=col_min,
        end_col=col_max,
        confidence=confidence,
        reason=reason,
        has_border=True,
    )


def _detect_table_regions(ws_openpyxl, max_row: int, max_col: int,
                          html_rows: Optional[List[List[CellData]]] = None
                          ) -> List[TableRegion]:
    """
    检测 Excel 中的表格区域，返回带置信度分级的 TableRegion 列表。

    检测策略（分层）：
    1. 边框检测：优先 openpyxl 原生，HTML 通道作为回退
    2. 列对齐检测（多行在相同列位置有内容 → 中/高置信）
    3. 两种策略的结果合并去重

    置信度分级：
    - HIGH:    有完整边框 + 3行以上 + 无 gap，或 列对齐 + 高填充率 + 5行以上
    - MEDIUM:  有边框但 gap>1 / 部分行无边框 / 列对齐但行数少
    - SUSPECT: 无边框但列对齐检测到疑似表格结构（2-3行）
    """
    regions = []

    # ── 策略1: 边框检测（openpyxl 原生，含线框图识别启发式）──
    border_regions = _detect_border_tables_openpyxl(ws_openpyxl, max_row, max_col)
    if border_regions:
        regions.extend(border_regions)

    # ── 策略2: 列对齐检测（基于 openpyxl 数据）──
    align_regions = _detect_aligned_tables(ws_openpyxl, max_row, max_col)

    # ── 合并两种策略的结果，去重 ──
    existing_rows = set()
    for reg in regions:
        for r in range(reg.start_row, reg.end_row + 1):
            existing_rows.add(r)

    for reg in align_regions:
        # 检查是否与已有区域重叠
        overlap = False
        for r in range(reg.start_row, reg.end_row + 1):
            if r in existing_rows:
                overlap = True
                break
        if not overlap:
            regions.append(reg)
        else:
            # 重叠时，如果对齐检测给出更高置信度，则升级
            # 但不覆盖线框图降级判定（reason 中包含"线框"标记）
            for existing in regions:
                if (existing.start_row <= reg.end_row and reg.start_row <= existing.end_row):
                    if reg.confidence == CONFIDENCE_HIGH and existing.confidence != CONFIDENCE_HIGH:
                        if '线框' not in existing.reason:
                            existing.confidence = CONFIDENCE_HIGH
                            existing.reason += f"; 列对齐验证升级"

    # 按起始行排序
    regions.sort(key=lambda r: r.start_row)
    return regions




def _detect_aligned_tables(ws_openpyxl, max_row: int, max_col: int) -> List[TableRegion]:
    """
    基于列对齐检测表格区域（不依赖边框，处理无边框表格）。
    核心思路：连续多行在相同列位置有内容 → 大概率是表格。
    """
    table_regions = []
    visited_rows = set()

    for r in range(1, max_row + 1):
        if r in visited_rows:
            continue

        # 统计该行有内容的列
        filled_cols = []
        for c in range(1, max_col + 1):
            cell = ws_openpyxl.cell(row=r, column=c)
            if cell.value is not None:
                filled_cols.append(c)

        if len(filled_cols) < 2:
            continue

        col_start = min(filled_cols)
        col_end = max(filled_cols)
        col_span = col_end - col_start + 1

        # 向下探测连续数据行（允许最多 1 行空行 gap）
        table_end = r
        consecutive_empty = 0
        for nr in range(r + 1, min(r + 500, max_row + 1)):
            nr_filled = []
            for c in range(col_start, col_end + 1):
                cell = ws_openpyxl.cell(row=nr, column=c)
                if cell.value is not None:
                    nr_filled.append(c)
            fill_rate = len(nr_filled) / col_span if col_span > 0 else 0
            if fill_rate >= 0.5:
                table_end = nr
                consecutive_empty = 0
            elif fill_rate > 0:
                # 部分填充，可能是 gap 行
                consecutive_empty += 1
                if consecutive_empty <= 1:
                    table_end = nr
                else:
                    break
            else:
                consecutive_empty += 1
                if consecutive_empty > 1:
                    break

        row_count = table_end - r + 1

        # 计算平均填充率
        total_filled = 0
        total_cells = 0
        for check_r in range(r, table_end + 1):
            for c in range(col_start, col_end + 1):
                total_cells += 1
                if ws_openpyxl.cell(row=check_r, column=c).value is not None:
                    total_filled += 1
        avg_fill = total_filled / total_cells if total_cells > 0 else 0

        # 置信度分级
        if row_count >= 5 and col_span >= 3 and avg_fill >= 0.6:
            confidence = CONFIDENCE_HIGH
            reason = f"列对齐({row_count}行×{col_span}列, 填充{avg_fill:.0%})"
        elif row_count >= 3 and col_span >= 2 and avg_fill >= 0.5:
            confidence = CONFIDENCE_MEDIUM
            reason = f"列对齐({row_count}行×{col_span}列, 填充{avg_fill:.0%})"
        elif row_count == 2 and col_span >= 2 and avg_fill >= 0.7:
            confidence = CONFIDENCE_SUSPECT
            reason = f"疑似表格({row_count}行×{col_span}列, 填充{avg_fill:.0%})"
        else:
            continue  # 不符合任何条件

        table_regions.append(TableRegion(
            start_row=r, end_row=table_end,
            start_col=col_start, end_col=col_end,
            confidence=confidence, reason=reason,
        ))
        for tr in range(r, table_end + 1):
            visited_rows.add(tr)

    # 合并相邻/重叠区域
    table_regions.sort(key=lambda r: r.start_row)
    merged = []
    for reg in table_regions:
        if merged and reg.start_row <= merged[-1].end_row + 2:
            prev = merged[-1]
            prev.end_row = max(prev.end_row, reg.end_row)
            prev.start_col = min(prev.start_col, reg.start_col)
            prev.end_col = max(prev.end_col, reg.end_col)
            # 合并后置信度取较低的
            if reg.confidence == CONFIDENCE_SUSPECT or prev.confidence == CONFIDENCE_SUSPECT:
                prev.confidence = CONFIDENCE_MEDIUM
            prev.reason += f" + 合并({reg.reason})"
        else:
            merged.append(reg)

    return merged


def _is_in_table_region(row: int, table_regions: List[TableRegion]) -> Optional[TableRegion]:
    """检查某行是否在数据表格区域内，返回区域或 None"""
    for region in table_regions:
        if region.start_row <= row <= region.end_row:
            return region
    return None


def _export_outside_table_text(ws_openpyxl, region, actual_max_col: int,
                                html_rows, hyperlinks_map: dict,
                                img_row_index: dict, min_content_col: int,
                                get_openpyxl_style_func) -> List[str]:
    """
    导出表格区域同行但在表格列范围之外的文本内容。
    解决策划文档中表格右侧/左侧有说明文字或嵌入图片被遗漏的问题。
    """
    sr, er = region.start_row, region.end_row
    sc, ec = region.start_col, region.end_col
    lines = []
    has_content = False

    for r in range(sr, er + 1):
        # 检查表格列范围外的图片
        imgs_here = img_row_index.get(r)
        if imgs_here:
            for img in imgs_here:
                # 只处理不在表格列范围内的图片
                img_col = img.get('start_col', 0)
                if img_col > ec or img_col < sc:
                    lines.append(f"📷 *嵌入图片 ({img['anchor']}，"
                                 f"{img['width']}x{img['height']}，见截图)*")
                    has_content = True

        # 收集表格列范围之外的非空单元格
        cells_in_row = []
        for c in range(1, actual_max_col + 1):
            if sc <= c <= ec:
                continue  # 跳过表格列范围内的单元格

            # 优先使用 HTML 通道数据
            html_cell = _get_html_cell(html_rows, r, c) if html_rows else None

            if html_cell and not html_cell.is_merged_slave and html_cell.text.strip():
                formatted = _format_text_with_style(
                    html_cell.text, html_cell.style, html_cell.html_content
                )
                if formatted:
                    link_target = hyperlinks_map.get((r, c))
                    if link_target:
                        formatted = f'[{formatted}]({link_target})'
                    cells_in_row.append((c, formatted))
            else:
                text = _get_cell_text(ws_openpyxl, r, c)
                if text is not None:
                    cell_obj = ws_openpyxl.cell(row=r, column=c)
                    style = get_openpyxl_style_func(cell_obj)
                    formatted = _format_text_with_style(text, style)
                    link_target = hyperlinks_map.get((r, c))
                    if link_target:
                        formatted = f'[{formatted}]({link_target})'
                    cells_in_row.append((c, formatted))

        if cells_in_row:
            if not has_content:
                # 第一次发现表格外内容，添加分隔标记
                lines.append("")
                lines.append(f"<!-- 表格区域(列{sc}-{ec})之外的同行内容 -->")
                lines.append("")
                has_content = True

            first_col = cells_in_row[0][0]
            indent_level = max(0, first_col - min_content_col)
            indent = "  " * indent_level
            combined_text = " ".join(text for _, text in cells_in_row)
            combined_text = combined_text.replace("\n", " ↵ ")
            if len(combined_text) > 500:
                combined_text = combined_text[:500] + "..."
            lines.append(f"R{r}| {indent}{combined_text}")

    return lines


def _export_table_region_md(ws_openpyxl, region: TableRegion, merged: dict,
                            html_rows: Optional[List[List[CellData]]] = None,
                            hyperlinks_map: Optional[Dict[Tuple[int, int], str]] = None) -> List[str]:
    """
    将一个数据表格区域导出为 Markdown 表格。
    如果有 HTML 通道数据，使用富文本格式化。
    """
    if hyperlinks_map is None:
        hyperlinks_map = {}
    sr, er = region.start_row, region.end_row
    sc, ec = region.start_col, region.end_col
    lines = []
    lines.append("")  # 空行分隔

    # 置信度标注
    if region.confidence == CONFIDENCE_MEDIUM:
        lines.append(f"> ⚠️ 低置信表格（{region.reason}）— 请 Agent 复核边界")
    elif region.confidence == CONFIDENCE_SUSPECT:
        lines.append(f"> 🟡 疑似表格（{region.reason}）— 请 Agent 判断是否为表格")

    for r in range(sr, er + 1):
        row_cells = []
        for c in range(sc, ec + 1):
            # 优先使用 HTML 通道的富文本数据
            html_cell = _get_html_cell(html_rows, r, c) if html_rows else None

            if html_cell and not html_cell.is_merged_slave:
                display = _format_text_with_style(
                    html_cell.text, html_cell.style, html_cell.html_content
                )
                # 表格中不能有换行
                display = display.replace('\n', ' ↵ ')
            else:
                # 回退到 openpyxl 数据
                cell = ws_openpyxl.cell(row=r, column=c)
                val = cell.value
                if val is None:
                    display = ""
                elif isinstance(val, (int, float)):
                    display = str(val)
                else:
                    display = str(val).replace("|", "\\|").replace("\n", " ↵ ")
                    if len(display) > 200:
                        display = display[:200] + "..."

            # 超链接包裹（在管道符转义之前）
            link_target = hyperlinks_map.get((r, c))
            if link_target and display.strip():
                display = f'[{display}]({link_target})'

            # 管道符转义
            display = display.replace("|", "\\|")

            merge_info = merged.get((r, c))
            if merge_info:
                display = f"{display} 【{merge_info}】"
            row_cells.append(display)

        # 跳过全空行
        if not any(c.strip() for c in row_cells):
            continue

        lines.append("| " + " | ".join(row_cells) + " |")
        # 在第一行（表头）后插入分隔线
        if r == sr:
            lines.append("|" + "|".join(["----"] * (ec - sc + 1)) + "|")

    lines.append("")  # 空行分隔
    return lines


def _get_html_cell(html_rows: Optional[List[List[CellData]]],
                   row_1based: int, col_1based: int) -> Optional[CellData]:
    """从 HTML 通道数据获取单元格（坐标转换：1-based → 0-based）"""
    if not html_rows:
        return None
    r_idx = row_1based - 1
    c_idx = col_1based - 1
    if 0 <= r_idx < len(html_rows):
        row = html_rows[r_idx]
        if 0 <= c_idx < len(row):
            return row[c_idx]
    return None


def _get_cell_text(ws_openpyxl, row: int, col: int) -> Optional[str]:
    """获取单元格文本，None 或空字符串返回 None"""
    cell = ws_openpyxl.cell(row=row, column=col)
    val = cell.value
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _export_sheet_markdown(ws_openpyxl, output_dir: Path, sheet_name: str,
                           screenshots: List[str], images_info: List[dict],
                           html_rows: Optional[List[List[CellData]]] = None,
                           comments_info: Optional[List[dict]] = None,
                           cf_info: Optional[List[dict]] = None,
                           hyperlinks_map: Optional[Dict[Tuple[int, int], str]] = None) -> str:
    """
    将工作表内容导出为 Markdown 文件。
    三通道融合输出：
    1. 自动表格检测（规则先行）→ 高置信区域直接生成 Markdown 表格
    2. 非表格区域 → 缩进文本树（保留富文本格式）
    3. 嵌入图片 → 内联标记
    4. 低置信/疑似区域 → 标记请 Agent 复核
    5. HTML 通道富文本格式（加粗/斜体/删除线/颜色等）
    6. 批注（comment）和条件格式（conditional_formatting）信息
    7. 超链接保留（外部 URL 和内部 Sheet 引用）
    """
    if hyperlinks_map is None:
        hyperlinks_map = {}
    lines = []
    lines.append(f"# Sheet: {sheet_name}\n")

    # ── 截图引用 ──
    if screenshots:
        lines.append("## 截图预览\n")
        for s in screenshots:
            lines.append(f"![{sheet_name}]({s})\n")
        lines.append("")

    # ── 批注信息 ──
    if comments_info:
        lines.append("## 💬 批注\n")
        for c in comments_info:
            author_str = f" ({c['author']})" if c['author'] else ""
            # 截断过长的批注文本
            text = c['text']
            if len(text) > 300:
                text = text[:300] + "..."
            text_oneline = text.replace('\n', ' ↵ ')
            lines.append(f"- **{c['cell']}**{author_str}: {text_oneline}")
        lines.append("")

    # ── 条件格式信息 ──
    if cf_info:
        lines.append("## 🎨 条件格式\n")
        for cf in cf_info:
            lines.append(f"- `{cf['range']}`: {cf['description']}")
        lines.append("")

    # ── 超链接信息 ──
    if hyperlinks_map:
        lines.append("## 🔗 超链接\n")
        for (r, c), target in sorted(hyperlinks_map.items()):
            cell_ref = f"{get_column_letter(c)}{r}"
            cell_text = ""
            try:
                val = ws_openpyxl.cell(row=r, column=c).value
                if val is not None:
                    cell_text = f" \"{str(val).strip()[:60]}\""
            except Exception:
                pass
            if target.startswith('#'):
                lines.append(f"- `{cell_ref}`{cell_text} → 内部引用: `{target[1:]}`")
            else:
                lines.append(f"- `{cell_ref}`{cell_text} → {target}")
        lines.append("")

    # ── 获取有效范围 ──
    if ws_openpyxl.max_row is None or ws_openpyxl.max_column is None:
        lines.append("> 空白工作表\n")
        md_path = output_dir / "data.md"
        md_path.write_text("\n".join(lines), encoding="utf-8")
        return "data.md"

    max_row = min(ws_openpyxl.max_row, 2000)
    max_col = min(ws_openpyxl.max_column, 200)

    # 找到实际有内容的范围
    actual_max_row = 0
    actual_max_col = 0
    row_has_content = set()
    for row in ws_openpyxl.iter_rows(min_row=1, max_row=max_row,
                                      min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)
                row_has_content.add(cell.row)

    if actual_max_row == 0:
        lines.append("> 无数据内容\n")
        md_path = output_dir / "data.md"
        md_path.write_text("\n".join(lines), encoding="utf-8")
        return "data.md"

    # ── 合并信息 ──
    merged = _get_merged_ranges(ws_openpyxl)

    # ── 图片行索引 ──
    img_row_index = _build_image_row_index(images_info)
    img_row_ranges = []
    for img in images_info:
        sr = img.get('start_row')
        er = img.get('end_row')
        if sr is not None and er is not None:
            img_row_ranges.append((sr, er, img))

    def _row_in_image_region(r: int) -> Optional[dict]:
        for sr, er, info in img_row_ranges:
            if sr <= r <= er:
                return info
        return None

    # ── 计算"自由文本"的最小起始列，作为缩进基准 ──
    min_content_col = actual_max_col
    for r in range(1, actual_max_row + 1):
        for c in range(1, actual_max_col + 1):
            if _get_cell_text(ws_openpyxl, r, c) is not None:
                min_content_col = min(min_content_col, c)
                break

    # ── 从 openpyxl 构建单元格富文本样式（当 HTML 通道不可用时的回退）──
    def _get_openpyxl_style(cell) -> CellStyle:
        """从 openpyxl cell 对象提取样式到 CellStyle"""
        style = CellStyle()
        try:
            font = cell.font
            if font:
                if font.bold:
                    style.bold = True
                if font.italic:
                    style.italic = True
                if font.strikethrough:
                    style.strikethrough = True
                if font.underline and font.underline != 'none':
                    style.underline = True
                color_hex = _color_to_hex(font.color)
                if color_hex and color_hex not in ('#000000', 'theme(1)', None):
                    if color_hex.startswith('#'):
                        normalized = _normalize_color(color_hex)
                        if normalized:
                            style.color = normalized
                if font.size and font.size != 11:
                    style.font_size = font.size
            fill = cell.fill
            if fill and fill.fill_type and fill.fill_type != 'none':
                fg_hex = _color_to_hex(fill.fgColor)
                if fg_hex and fg_hex.startswith('#'):
                    normalized = _normalize_bg_color(fg_hex)
                    if normalized:
                        style.bg_color = normalized
        except Exception:
            pass
        return style

    # ── 表格区域检测（规则先行）──
    table_regions = _detect_table_regions(ws_openpyxl, actual_max_row, actual_max_col, html_rows)

    if table_regions:
        high_count = sum(1 for t in table_regions if t.confidence == CONFIDENCE_HIGH)
        med_count = sum(1 for t in table_regions if t.confidence == CONFIDENCE_MEDIUM)
        sus_count = sum(1 for t in table_regions if t.confidence == CONFIDENCE_SUSPECT)
        lines.append(f"## 表格检测结果\n")
        lines.append(f"共检测到 {len(table_regions)} 个表格区域"
                     f"（✅高置信 {high_count}, ⚠️中置信 {med_count}, 🟡疑似 {sus_count}）\n")
        for i, reg in enumerate(table_regions, 1):
            conf_icon = {"high": "✅", "medium": "⚠️", "suspect": "🟡"}[reg.confidence]
            lines.append(f"- 区域{i}: 行{reg.start_row}-{reg.end_row}, "
                         f"列{reg.start_col}-{reg.end_col} {conf_icon} {reg.reason}")
        lines.append("")

    # 构建表格行集合用于快速查找
    table_row_set = set()
    for reg in table_regions:
        for r in range(reg.start_row, reg.end_row + 1):
            table_row_set.add(r)

    # ── 输出内容（表格检测 + 富文本）──
    lines.append("## 内容\n")

    r = 1
    while r <= actual_max_row:
        # ─ 连续空行压缩 ─
        if r not in row_has_content and r not in table_row_set:
            empty_start = r
            while r <= actual_max_row and r not in row_has_content and r not in table_row_set:
                r += 1
            empty_end = r - 1
            empty_count = empty_end - empty_start + 1

            if empty_count > 2:
                img_in_region = _row_in_image_region(empty_start)
                if img_in_region:
                    lines.append(f"\n📷 *嵌入图片 (行{empty_start}-{empty_end}，{img_in_region['anchor']}，"
                                 f"{img_in_region['width']}x{img_in_region['height']}，见截图)*\n")
            elif empty_count > 0:
                img_in_region = _row_in_image_region(empty_start)
                if img_in_region:
                    lines.append(f"\n📷 *嵌入图片 ({img_in_region['anchor']}，"
                                 f"{img_in_region['width']}x{img_in_region['height']}，见截图)*\n")
            continue

        # ─ 检查是否在表格区域内 → 生成 Markdown 表格 ─
        table_reg = _is_in_table_region(r, table_regions)
        if table_reg:
            # 检查表格区域之前/内的图片
            for tr in range(table_reg.start_row, table_reg.end_row + 1):
                imgs_here = img_row_index.get(tr)
                if imgs_here:
                    for img in imgs_here:
                        lines.append(f"📷 *嵌入图片 ({img['anchor']}，"
                                     f"{img['width']}x{img['height']}，见截图)*")

            table_lines = _export_table_region_md(ws_openpyxl, table_reg, merged, html_rows, hyperlinks_map)
            lines.extend(table_lines)

            # ── 导出表格列范围之外的文本（同行但在表格区域列外的内容）──
            outside_lines = _export_outside_table_text(
                ws_openpyxl, table_reg, actual_max_col, html_rows, hyperlinks_map,
                img_row_index, min_content_col, _get_openpyxl_style
            )
            if outside_lines:
                lines.extend(outside_lines)

            r = table_reg.end_row + 1
            continue

        # ─ 非表格区域：自由文本行（用缩进层级表示）─
        # 检查图片
        imgs_here = img_row_index.get(r)
        if imgs_here:
            for img in imgs_here:
                lines.append(f"📷 *嵌入图片 ({img['anchor']}，"
                             f"{img['width']}x{img['height']}，见截图)*")

        # 收集该行所有非空单元格（优先 HTML 通道富文本）
        cells_in_row = []
        for c in range(1, actual_max_col + 1):
            # 优先使用 HTML 通道数据
            html_cell = _get_html_cell(html_rows, r, c) if html_rows else None

            if html_cell and not html_cell.is_merged_slave and html_cell.text.strip():
                formatted = _format_text_with_style(
                    html_cell.text, html_cell.style, html_cell.html_content
                )
                if formatted:
                    # 超链接包裹
                    link_target = hyperlinks_map.get((r, c))
                    if link_target:
                        formatted = f'[{formatted}]({link_target})'
                    cells_in_row.append((c, formatted))
            else:
                # 回退到 openpyxl
                text = _get_cell_text(ws_openpyxl, r, c)
                if text is not None:
                    cell_obj = ws_openpyxl.cell(row=r, column=c)
                    style = _get_openpyxl_style(cell_obj)
                    formatted = _format_text_with_style(text, style)
                    # 超链接包裹
                    link_target = hyperlinks_map.get((r, c))
                    if link_target:
                        formatted = f'[{formatted}]({link_target})'
                    cells_in_row.append((c, formatted))

        if cells_in_row:
            first_col = cells_in_row[0][0]
            indent_level = max(0, first_col - min_content_col)
            indent = "  " * indent_level

            combined_text = " ".join(text for _, text in cells_in_row)
            combined_text = combined_text.replace("\n", " ↵ ")
            if len(combined_text) > 500:
                combined_text = combined_text[:500] + "..."

            lines.append(f"R{r}| {indent}{combined_text}")

        r += 1

    lines.append("")

    # ── 写入文件 ──
    md_path = output_dir / "data.md"
    md_path.write_text("\n".join(lines), encoding="utf-8")
    size_kb = md_path.stat().st_size / 1024
    table_info = f", {len(table_regions)}个表格" if table_regions else ""
    html_info = ", HTML富文本" if html_rows else ""
    print(f"    [文本] 已保存: data.md ({size_kb:.1f} KB, {actual_max_row}行x{actual_max_col}列{table_info}{html_info})")
    return "data.md"


def _get_comments_info(ws_openpyxl, max_row: int, max_col: int) -> List[dict]:
    """
    获取工作表中所有批注（comment）信息。
    返回 [{"cell": "B5", "row": 5, "col": 2, "author": "张三", "text": "..."}]
    """
    comments = []
    for row in ws_openpyxl.iter_rows(min_row=1, max_row=max_row,
                                      min_col=1, max_col=max_col):
        for cell in row:
            if cell.comment:
                comment_text = str(cell.comment.text).strip() if cell.comment.text else ""
                author = str(cell.comment.author).strip() if cell.comment.author else ""
                if comment_text:
                    comments.append({
                        'cell': f"{get_column_letter(cell.column)}{cell.row}",
                        'row': cell.row,
                        'col': cell.column,
                        'author': author,
                        'text': comment_text,
                    })
    return comments


def _get_conditional_formatting_info(ws_openpyxl) -> List[dict]:
    """
    获取工作表中的条件格式规则信息。
    返回 [{"range": "A1:A100", "type": "dataBar", "description": "数据条 ..."}]
    """
    cf_info = []
    for cf in ws_openpyxl.conditional_formatting:
        ranges_str = str(cf)
        for rule in cf.rules:
            rule_type = rule.type if hasattr(rule, 'type') else 'unknown'
            description = ""

            if rule_type == 'dataBar':
                description = "数据条（数值越大条越长）"
            elif rule_type == 'colorScale':
                description = "色阶（数值映射到颜色渐变）"
            elif rule_type == 'iconSet':
                description = "图标集（数值映射到图标）"
            elif rule_type == 'cellIs':
                op = getattr(rule, 'operator', '')
                formula = getattr(rule, 'formula', [])
                formula_str = ', '.join(str(f) for f in formula) if formula else ''
                description = f"单元格值 {op} {formula_str}"
            elif rule_type == 'containsText':
                text = getattr(rule, 'text', '')
                description = f"包含文本 '{text}'"
            elif rule_type == 'duplicateValues':
                description = "重复值高亮"
            elif rule_type == 'expression':
                formula = getattr(rule, 'formula', [])
                formula_str = ', '.join(str(f) for f in formula) if formula else ''
                description = f"公式: {formula_str}"
            else:
                description = f"类型: {rule_type}"

            # 提取格式效果
            effects = []
            if hasattr(rule, 'dxf') and rule.dxf:
                dxf = rule.dxf
                if dxf.font:
                    if dxf.font.bold:
                        effects.append("加粗")
                    if dxf.font.color and hasattr(dxf.font.color, 'rgb') and dxf.font.color.rgb:
                        effects.append(f"字色=#{str(dxf.font.color.rgb)[2:]}")
                if dxf.fill and dxf.fill.fgColor and hasattr(dxf.fill.fgColor, 'rgb') and dxf.fill.fgColor.rgb:
                    effects.append(f"背景=#{str(dxf.fill.fgColor.rgb)[2:]}")

            if effects:
                description += f" → {', '.join(effects)}"

            cf_info.append({
                'range': ranges_str,
                'type': rule_type,
                'description': description,
            })
    return cf_info


def _get_hyperlinks_map(ws_openpyxl) -> Dict[Tuple[int, int], str]:
    """
    获取工作表中所有超链接，返回 {(row, col): target_url} 字典。
    支持外部链接（URL）和内部链接（Sheet 引用）。
    """
    links = {}
    # 方式1：遍历 ws.hyperlinks（openpyxl 存储的超链接列表）
    if hasattr(ws_openpyxl, 'hyperlinks'):
        for hl in ws_openpyxl.hyperlinks:
            try:
                ref = hl.ref
                # ref 可能是单单元格 "A1" 或范围 "A1:B2"，取左上角
                if ':' in ref:
                    ref = ref.split(':')[0]
                from openpyxl.utils import coordinate_to_tuple
                row, col = coordinate_to_tuple(ref)
                target = hl.target  # 外部 URL
                if target:
                    links[(row, col)] = target
                elif hl.location:
                    # 内部链接（如 "Sheet2!A1"）
                    links[(row, col)] = f"#{hl.location}"
            except Exception:
                continue

    # 方式2：逐单元格检查 cell.hyperlink（覆盖方式1可能遗漏的情况）
    max_row = min(ws_openpyxl.max_row or 0, 2000)
    max_col = min(ws_openpyxl.max_column or 0, 200)
    if max_row > 0 and max_col > 0:
        for row in ws_openpyxl.iter_rows(min_row=1, max_row=max_row,
                                          min_col=1, max_col=max_col):
            for cell in row:
                if cell.hyperlink and (cell.row, cell.column) not in links:
                    target = cell.hyperlink.target
                    if target:
                        links[(cell.row, cell.column)] = target
                    elif cell.hyperlink.location:
                        links[(cell.row, cell.column)] = f"#{cell.hyperlink.location}"
    return links


def _get_images_info(ws_openpyxl) -> List[dict]:
    """
    获取工作表中嵌入图片的位置信息。
    返回包含 anchor, width, height, start_row, end_row, start_col, end_col 的字典列表。

    对于 OneCellAnchor（仅有起始锚点、无 to 锚点）的图片，
    根据图片的 width/height（EMU 或像素）和列宽/行高推算 end_col/end_row，
    避免截图时列/行范围不足导致图片被裁切。
    """
    images = []
    for img in ws_openpyxl._images:
        anchor_str = '未知位置'
        start_row = None
        end_row = None
        has_to_anchor = False
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from is not None:
                fr = anchor._from
                start_row = fr.row + 1  # openpyxl 0-based → 1-based
                anchor_str = f"{get_column_letter(fr.col + 1)}{fr.row + 1}"
                if hasattr(anchor, 'to') and anchor.to is not None:
                    to = anchor.to
                    end_row = to.row + 1
                    has_to_anchor = True
                    anchor_str += f":{get_column_letter(to.col + 1)}{to.row + 1}"
                else:
                    end_row = start_row
            elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                start_row = anchor.row + 1
                end_row = start_row
                anchor_str = f"{get_column_letter(anchor.col + 1)}{anchor.row + 1}"
        except Exception:
            pass

        # 提取列信息
        start_col = None
        end_col = None
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from is not None:
                start_col = anchor._from.col + 1  # 0-based → 1-based
                if hasattr(anchor, 'to') and anchor.to is not None:
                    end_col = anchor.to.col + 1
                else:
                    end_col = start_col
            elif hasattr(anchor, 'col'):
                start_col = anchor.col + 1
                end_col = start_col
        except Exception:
            pass

        # ── OneCellAnchor：存储原始 EMU 尺寸，延迟到 COM 阶段精确计算 end_col/end_row ──
        # openpyxl img.width/height 单位为 EMU (English Metric Units)
        # 1 EMU = 1/914400 inch; 1pt = 12700 EMU
        # 精确的列宽/行高只能从 COM 获取（ws_com.Columns(col).Width 返回磅值），
        # openpyxl 的字符单位列宽转像素依赖字体，存在不可控误差。
        width_emu = None
        height_emu = None
        if not has_to_anchor and start_col is not None and start_row is not None:
            img_w = img.width if hasattr(img, 'width') else None
            img_h = img.height if hasattr(img, 'height') else None
            if img_w is not None and isinstance(img_w, (int, float)) and img_w > 0:
                width_emu = float(img_w)
            if img_h is not None and isinstance(img_h, (int, float)) and img_h > 0:
                height_emu = float(img_h)

        info = {
            'anchor': anchor_str,
            'width': img.width if hasattr(img, 'width') else '?',
            'height': img.height if hasattr(img, 'height') else '?',
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col,
            'width_emu': width_emu,    # OneCellAnchor 图片的原始宽度 (EMU)，供 COM 阶段精确计算
            'height_emu': height_emu,  # OneCellAnchor 图片的原始高度 (EMU)，供 COM 阶段精确计算
        }
        images.append(info)
    return images


def _refine_image_bounds_com(ws_com, images_info: List[dict]):
    """
    COM 阶段精确计算 OneCellAnchor 图片的 end_col / end_row。

    原理：
    - ws_com.Columns(col).Width 返回精确的列宽（磅 pt）
    - ws_com.Rows(row).Height 返回精确的行高（磅 pt）
    - 图片 EMU → 磅：emu / 12700.0（精确定义，无经验常数）
    - 累加列宽/行高直到容纳图片尺寸，得到精确的 end_col / end_row

    修改 images_info 中对应项的 end_col / end_row / anchor（就地修改）。
    仅处理有 width_emu / height_emu 的图片（即 OneCellAnchor 类型）。
    """
    for img in images_info:
        w_emu = img.get('width_emu')
        h_emu = img.get('height_emu')
        start_col = img.get('start_col')
        start_row = img.get('start_row')

        if w_emu and start_col:
            # EMU → 磅（1 pt = 12700 EMU，精确定义）
            # EMU > 10000 → 真正的 EMU；否则可能是像素/磅，按磅处理
            if w_emu > 10000:
                img_w_pt = w_emu / 12700.0
            else:
                img_w_pt = w_emu  # 小值视为磅

            remaining = img_w_pt
            col = start_col
            while remaining > 0 and col <= 500:
                remaining -= ws_com.Columns(col).Width  # 精确磅值
                col += 1
            new_end_col = col - 1
            if new_end_col > (img.get('end_col') or 0):
                img['end_col'] = new_end_col
                # 更新 anchor 字符串
                old_anchor = img.get('anchor', '')
                prefix = old_anchor.split(':')[0] if ':' in old_anchor else old_anchor
                img['anchor'] = f"{prefix}:{get_column_letter(new_end_col)}{img.get('end_row', start_row)}"

        if h_emu and start_row:
            if h_emu > 10000:
                img_h_pt = h_emu / 12700.0
            else:
                img_h_pt = h_emu

            remaining = img_h_pt
            row = start_row
            while remaining > 0 and row <= 5000:
                remaining -= ws_com.Rows(row).Height  # 精确磅值
                row += 1
            new_end_row = row - 1
            if new_end_row > (img.get('end_row') or 0):
                img['end_row'] = new_end_row


# ──────────────────────────────────────────────
# 增量更新辅助函数
# ──────────────────────────────────────────────

def _load_export_manifest(output_dir: Path) -> dict:
    """读取导出清单文件，用于增量更新判断。"""
    import json as _json
    manifest_path = output_dir / ".export_manifest.json"
    if manifest_path.exists():
        try:
            with open(manifest_path, 'r', encoding='utf-8') as f:
                return _json.load(f)
        except Exception:
            pass
    return {}


def _save_export_manifest(output_dir: Path, manifest: dict):
    """保存导出清单文件。"""
    import json as _json
    manifest_path = output_dir / ".export_manifest.json"
    with open(manifest_path, 'w', encoding='utf-8') as f:
        _json.dump(manifest, f, ensure_ascii=False, indent=2)


def _calc_sheet_checksum(ws, images: list) -> str:
    """
    计算 sheet 内容的轻量级校验值，用于判断是否需要重新导出。
    基于 sheet 的维度信息、单元格内容和图片数量生成 MD5。
    """
    import hashlib
    h = hashlib.md5()
    # 维度信息
    h.update(f"dim:{ws.min_row}:{ws.max_row}:{ws.min_column}:{ws.max_column}".encode())
    # 图片数量和位置
    h.update(f"imgs:{len(images)}".encode())
    for img in images:
        h.update(f"img:{img.get('anchor', '')}".encode())
    # 单元格内容全量哈希
    max_row = min(ws.max_row or 0, 2000)
    max_col = min(ws.max_column or 0, 200)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                h.update(f"{cell.row},{cell.column}:{cell.value}".encode('utf-8', errors='replace'))
    return h.hexdigest()


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────

def export_excel(input_path: str, output_dir: Optional[str] = None,
                 selected_sheets: Optional[List[str]] = None,
                 tag: Optional[str] = None):
    """
    三通道导出 Excel：截图 + 结构化文本 + HTML富文本（支持增量更新）

    通道1: openpyxl 预读 → 精确文本 + 内容边界 + 校验值 + 批注 + 条件格式
    通道2: Excel COM → HTML → CSS/HTML解析 → 富文本样式 + 边框信息

    注意：截图不在此步骤执行。Agent 分析 data.md 后按需调用截图工具。

    通过 .export_manifest.json 记录每个 sheet 的内容校验值，
    再次导出时自动跳过内容未变化的 sheet。

    Args:
        input_path: 输入 Excel 文件路径
        output_dir: 输出目录（默认: 同目录下 {tag}_export/）
        selected_sheets: 要导出的 sheet 名称列表（默认 None 导出全部）
        tag: 输出目录标签（ASCII），用于拼接 {tag}_export 目录名。
             如果提供了 output_dir 则忽略 tag。
    """
    input_path = Path(input_path).resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"文件不存在: {input_path}")

    if output_dir is None:
        dir_tag = tag if tag else input_path.stem
        output_dir = input_path.parent / f"{dir_tag}_export"
    else:
        output_dir = Path(output_dir).resolve()

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] 输入文件: {input_path}")
    print(f"[INFO] 输出目录: {output_dir}")

    # ── 增量更新: 加载导出清单 ──
    manifest = _load_export_manifest(output_dir)

    # ── 第一步: 用 openpyxl 预读文本数据（通道2）──
    print("\n[STEP 1] 读取结构化数据 (openpyxl) ...")
    wb_openpyxl = load_workbook(str(input_path), data_only=True, rich_text=False)
    all_sheet_names = wb_openpyxl.sheetnames
    print(f"  共 {len(all_sheet_names)} 个工作表: {', '.join(all_sheet_names)}")

    # 过滤选中的 sheet
    if selected_sheets:
        sheet_names = [s for s in all_sheet_names if s in selected_sheets]
        skipped = [s for s in all_sheet_names if s not in selected_sheets]
        if skipped:
            print(f"  跳过未选中的工作表: {', '.join(skipped)}")
        print(f"  将导出 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")
    else:
        sheet_names = all_sheet_names

    # 建立 sheet_name → sheet_NN 索引映射（基于在 all_sheet_names 中的位置）
    sheet_index_map = {}  # {sheet_name: index}
    sheet_dir_map = {}    # {sheet_name: "sheet_NN"}
    for i, sn in enumerate(all_sheet_names, 1):
        sheet_index_map[sn] = i
        sheet_dir_map[sn] = _sheet_dirname(i)

    # 预存 openpyxl 数据，并计算每个 sheet 的内容边界和校验值
    openpyxl_data = {}
    new_checksums = {}
    sheets_to_export = []
    sheets_unchanged = []

    for name in sheet_names:
        ws = wb_openpyxl[name]
        images = _get_images_info(ws)
        bounds = _calc_content_bounds(ws, images)
        checksum = _calc_sheet_checksum(ws, images)
        openpyxl_data[name] = {
            'ws': ws,
            'images': images,
            'content_bounds': bounds,
        }
        new_checksums[name] = checksum

        # 增量判断：比较校验值 + 检查 data.md 是否存在
        old_checksum = manifest.get('sheets', {}).get(name, {}).get('checksum')
        sheet_dir = output_dir / sheet_dir_map[name]
        has_data_md = (sheet_dir / 'data.md').exists()

        if old_checksum == checksum and has_data_md:
            sheets_unchanged.append(name)
            print(f"  [{name}] 内容未变化，跳过重新导出 ✓")
        else:
            sheets_to_export.append(name)
            img_count = len(images)
            if img_count > 0:
                print(f"  [{name}] {img_count} 张嵌入图片")
            if bounds:
                print(f"  [{name}] 内容边界: 行{bounds[0]}-{bounds[1]}, 列{bounds[2]}-{bounds[3]}")
            if old_checksum and old_checksum != checksum:
                print(f"  [{name}] 内容已变化，将重新导出 ↻")

    if sheets_unchanged:
        print(f"\n  [增量] {len(sheets_unchanged)} 个 sheet 无变化: {', '.join(sheets_unchanged)}")
    if sheets_to_export:
        print(f"  [增量] {len(sheets_to_export)} 个 sheet 需导出: {', '.join(sheets_to_export)}")
    else:
        print(f"\n[OK] 所有选中的 sheet 均无变化，无需重新导出。")
        print(f"[OK] 已有导出结果在: {output_dir}")
        wb_openpyxl.close()
        return

    # ── 第一步(b): 提取批注、条件格式和超链接 ──
    print("\n  [增强] 提取批注、条件格式和超链接 ...")
    comments_map = {}      # {sheet_name: [comment_info]}
    cf_map = {}            # {sheet_name: [cf_info]}
    hyperlinks_map_all = {}  # {sheet_name: {(row, col): target}}
    for name in sheets_to_export:
        ws = openpyxl_data[name]['ws']
        bounds = openpyxl_data[name].get('content_bounds')
        max_row = min(ws.max_row or 0, 2000)
        max_col = min(ws.max_column or 0, 200)

        comments = _get_comments_info(ws, max_row, max_col)
        if comments:
            comments_map[name] = comments
            print(f"  [{name}] {len(comments)} 条批注")

        cf = _get_conditional_formatting_info(ws)
        if cf:
            cf_map[name] = cf
            print(f"  [{name}] {len(cf)} 条条件格式规则")

        hyperlinks = _get_hyperlinks_map(ws)
        if hyperlinks:
            hyperlinks_map_all[name] = hyperlinks
            print(f"  [{name}] {len(hyperlinks)} 个超链接")

    # ── 第二步: COM 会话 — HTML导出（通道2）──
    # 注意：截图不在此处执行，由 Agent 分析 data.md 后按需调用截图工具
    print("\n[STEP 2] Excel COM HTML 导出 ...")
    pythoncom.CoInitialize()
    excel = None
    wb_com = None
    html_channel_data = {}  # {sheet_name: List[List[CellData]]}

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        wb_com = excel.Workbooks.Open(str(input_path), ReadOnly=True)

        # ── HTML 导出 ──
        print("\n  [HTML] 导出 HTML 格式 ...")
        try:
            # 构建 content_bounds 映射，用于裁剪 HTML 导出范围（性能优化）
            bounds_map = {}
            for sname in sheets_to_export:
                b = openpyxl_data[sname].get('content_bounds')
                if b:
                    bounds_map[sname] = b
            html_paths = _export_sheets_to_html(excel, wb_com, output_dir, sheets_to_export,
                                                bounds_map=bounds_map)

            # 解析 HTML → CellData
            for name, htm_path in html_paths.items():
                try:
                    css_classes, rows_data = _load_html_channel_data(htm_path)
                    if rows_data:
                        html_channel_data[name] = rows_data
                        print(f"  [{name}] HTML 解析完成: {len(rows_data)}行, "
                              f"{len(css_classes)} 个CSS样式类")
                except Exception as e:
                    print(f"  [{name}] HTML 解析失败: {e}")
        except Exception as e:
            print(f"  [HTML] HTML 导出失败（降级为双通道模式）: {e}")

    except Exception as e:
        print(f"[ERROR] COM 操作失败: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            if wb_com:
                wb_com.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        wb_com = None
        excel = None
        pythoncom.CoUninitialize()
        print("  [COM] Excel 进程已关闭")

    # ── 第三步: 生成 Markdown（多通道融合）──
    print("\n[STEP 3] 生成结构化 Markdown（多通道融合）...")
    for name in sheets_to_export:
        sheet_dir = output_dir / sheet_dir_map[name]
        sheet_dir.mkdir(parents=True, exist_ok=True)

        ws = openpyxl_data[name]['ws']
        images = openpyxl_data[name]['images']
        screenshots = []  # 截图由 Agent 后续按需生成
        html_rows = html_channel_data.get(name)  # 可能为 None（降级）
        comments = comments_map.get(name)
        cf = cf_map.get(name)

        hyperlinks = hyperlinks_map_all.get(name)
        _export_sheet_markdown(ws, sheet_dir, name, screenshots, images, html_rows,
                               comments_info=comments, cf_info=cf,
                               hyperlinks_map=hyperlinks)

    # ── 清理临时 HTML 文件 ──
    tmp_html_dir = output_dir / "_tmp_html"
    if tmp_html_dir.exists():
        try:
            shutil.rmtree(tmp_html_dir, ignore_errors=True)
            print("  [清理] 临时 HTML 文件已删除")
        except Exception:
            pass

    # ── 保存导出清单（更新所有 sheet 的校验值） ──
    manifest_sheets = manifest.get('sheets', {})
    for name in sheet_names:
        manifest_sheets[name] = {
            'checksum': new_checksums[name],
            'exported_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        }
    manifest['sheets'] = manifest_sheets
    manifest['excel_path'] = str(input_path)
    manifest['last_updated'] = time.strftime('%Y-%m-%d %H:%M:%S')
    _save_export_manifest(output_dir, manifest)

    # 生成 sheet 映射文件，供 Agent 查找 sheet_NN ↔ 原始名称
    import json as _json
    mapping = [
        {"dir": sheet_dir_map[name], "index": sheet_index_map[name], "name": name}
        for name in sheet_names
    ]
    mapping_path = output_dir / "sheet_mapping.json"
    mapping_path.write_text(
        _json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  已生成: sheet_mapping.json")

    wb_openpyxl.close()

    exported_count = len(sheets_to_export)
    cached_count = len(sheets_unchanged)
    html_count = len(html_channel_data)
    summary_parts = [f"{exported_count} 个 sheet 已导出"]
    if cached_count > 0:
        summary_parts.append(f"{cached_count} 个使用缓存")
    if html_count > 0:
        summary_parts.append(f"{html_count} 个含HTML富文本")
    comment_count = sum(len(v) for v in comments_map.values())
    cf_count = sum(len(v) for v in cf_map.values())
    link_count = sum(len(v) for v in hyperlinks_map_all.values())
    if comment_count > 0:
        summary_parts.append(f"{comment_count} 条批注")
    if cf_count > 0:
        summary_parts.append(f"{cf_count} 条条件格式")
    if link_count > 0:
        summary_parts.append(f"{link_count} 个超链接")
    print(f"\n{'='*60}")
    print(f"[OK] data.md 导出完成！{', '.join(summary_parts)}")
    print(f"[OK] 输出目录: {output_dir}")
    print(f"[OK] 截图未在此步骤生成，由 Agent 分析 data.md 后按需调用截图工具")
    print(f"{'='*60}")


def _safe_dirname(name: str) -> str:
    """将 Sheet 名转为安全的目录名（仅保留 ASCII）"""
    # 替换 Windows 文件名非法字符
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, '_')
    return name.strip()


def _sheet_dirname(index: int) -> str:
    """根据 sheet 索引生成数据目录名: sheet_01_data, sheet_02_data, ..."""
    return f"sheet_{index:02d}_data"


# ──────────────────────────────────────────────
# 公开 API — 供外部脚本/Skill 调用
# ──────────────────────────────────────────────

def get_sheet_info(excel_path: str) -> List[dict]:
    """
    返回 Excel 文件中所有 sheet 的基础信息。

    Args:
        excel_path: Excel 文件路径

    Returns:
        List[dict]，每个元素包含:
        - name: sheet 名称
        - index: 1-based 索引
        - content_bounds: (min_row, max_row, min_col, max_col) 或 None
        - row_count: 内容行数（基于 content_bounds）
        - col_count: 内容列数
        - image_count: 嵌入图片数量
    """
    excel_path = Path(excel_path).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"文件不存在: {excel_path}")

    wb = load_workbook(str(excel_path), data_only=True, rich_text=False)
    result = []

    for idx, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        images = _get_images_info(ws)
        bounds = _calc_content_bounds(ws, images)

        info = {
            'name': name,
            'index': idx,
            'content_bounds': bounds,
            'row_count': (bounds[1] - bounds[0] + 1) if bounds else 0,
            'col_count': (bounds[3] - bounds[2] + 1) if bounds else 0,
            'image_count': len(images),
        }
        result.append(info)

    wb.close()
    return result


def export_sheet_range_screenshot(excel_path: str, sheet_name: str,
                                  ranges: List,
                                  output_dir: str) -> List[dict]:
    """
    对指定 sheet 的一组行范围进行截图，一次 COM 会话完成所有截图。

    Args:
        excel_path: Excel 文件路径
        sheet_name: 目标 sheet 名称
        ranges: 行范围列表，每个元素为:
                - (start_row, end_row, output_filename) 或
                - (start_row, end_row, output_filename, max_col)
                max_col 可选，由 Agent 指定的最大内容列号（脚本自动加安全距离）
        output_dir: 截图输出目录

    Returns:
        List[dict]，每个元素包含:
        - filename: 输出文件名
        - start_row: 起始行
        - end_row: 结束行
        - success: 是否成功
        - files: 生成的文件名列表（分片时可能有多个）
        - error: 错误信息（失败时）
    """
    excel_path = Path(excel_path).resolve()
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        raise FileNotFoundError(f"文件不存在: {excel_path}")

    # 先用 openpyxl 获取 content_bounds
    wb_openpyxl = load_workbook(str(excel_path), data_only=True, rich_text=False)
    if sheet_name not in wb_openpyxl.sheetnames:
        wb_openpyxl.close()
        raise ValueError(f"Sheet '{sheet_name}' 不存在")

    ws = wb_openpyxl[sheet_name]
    images = _get_images_info(ws)
    content_bounds = _calc_content_bounds(ws, images)

    if not content_bounds:
        wb_openpyxl.close()
        return [{'filename': r[2], 'start_row': r[0], 'end_row': r[1],
                 'success': False, 'files': [], 'error': 'Sheet 无内容'} for r in ranges]

    # 解析 ranges（支持带/不带 max_col 的格式）
    parsed_ranges = []
    for r in ranges:
        if len(r) >= 4:
            parsed_ranges.append((r[0], r[1], r[2], r[3]))  # (start, end, filename, max_col)
        else:
            parsed_ranges.append((r[0], r[1], r[2], None))  # (start, end, filename, None)

    # 为每个行范围预计算专属的列边界
    range_col_bounds = {}
    for start_row, end_row, filename, max_col in parsed_ranges:
        col_bounds = _calc_row_range_col_bounds(ws, start_row, end_row, images, max_col_override=max_col)
        if col_bounds:
            range_col_bounds[(start_row, end_row)] = col_bounds

    wb_openpyxl.close()

    # ── 第一阶段：用一个 COM 实例做精确校正（只读操作，不调用 CopyPicture）──
    results = []
    pythoncom.CoInitialize()

    try:
        excel_ref = win32com.client.DispatchEx("Excel.Application")
        excel_ref.Visible = False
        excel_ref.DisplayAlerts = False
        excel_ref.ScreenUpdating = False
        wb_ref = excel_ref.Workbooks.Open(str(excel_path), ReadOnly=True)
        ws_ref = wb_ref.Sheets(sheet_name)

        # COM 阶段：精确校正 OneCellAnchor 图片的 end_col / end_row
        _refine_image_bounds_com(ws_ref, images)

        wb_ref.Close(SaveChanges=False)
        excel_ref.Quit()
        del ws_ref, wb_ref, excel_ref
    except Exception as e:
        print(f"    [警告] COM refine 失败，使用 openpyxl 估算值: {e}")
        try:
            wb_ref.Close(SaveChanges=False)
        except:
            pass
        try:
            excel_ref.Quit()
        except:
            pass

    # 用精确的图片边界重算 content_bounds
    if content_bounds:
        cb_min_r, cb_max_r, cb_min_c, cb_max_c = content_bounds
        for img in images:
            ec = img.get('end_col')
            er = img.get('end_row')
            sc = img.get('start_col')
            sr = img.get('start_row')
            if ec is not None:
                cb_max_c = max(cb_max_c, ec)
            if er is not None:
                cb_max_r = max(cb_max_r, er)
            if sc is not None:
                cb_min_c = min(cb_min_c, sc)
            if sr is not None:
                cb_min_r = min(cb_min_r, sr)
        content_bounds = (cb_min_r, cb_max_r, cb_min_c, cb_max_c)

    # 用精确的图片边界重算每个行范围的列边界
    for key in list(range_col_bounds.keys()):
        rr_min_c, rr_max_c = range_col_bounds[key]
        rr_start, rr_end = key
        for img in images:
            sr = img.get('start_row', 0)
            er = img.get('end_row') or sr
            if sr <= rr_end and er >= rr_start:
                ec = img.get('end_col')
                sc = img.get('start_col')
                if ec is not None and ec + ROW_RANGE_COL_SAFETY > rr_max_c:
                    rr_max_c = ec + ROW_RANGE_COL_SAFETY
                if sc is not None and sc < rr_min_c:
                    rr_min_c = sc
        range_col_bounds[key] = (rr_min_c, rr_max_c)

    # ── 第二阶段：每个 zone 独立创建/销毁 COM 实例进行截图 ──
    # 关键：CopyPicture 会破坏 COM 会话状态，同一实例的后续调用可能
    # 全部失败（RPC_E_CALL_REJECTED）。因此每个 zone 必须用独立实例。
    for start_row, end_row, filename, max_col in parsed_ranges:
        excel_snap = None
        wb_snap = None
        try:
            excel_snap = win32com.client.DispatchEx("Excel.Application")
            excel_snap.Visible = False
            excel_snap.DisplayAlerts = False
            excel_snap.ScreenUpdating = False
            wb_snap = excel_snap.Workbooks.Open(str(excel_path), ReadOnly=True)
            ws_snap = wb_snap.Sheets(sheet_name)

            # 解除 sheet 工作组并激活目标 sheet，避免 CopyPicture 失败
            try:
                ws_snap.Activate()
                time.sleep(1.0)
            except Exception:
                pass

            # 使用行范围专属的列边界（比全 sheet 的更紧凑）
            col_bounds = range_col_bounds.get((start_row, end_row))
            if col_bounds:
                rr_min_c, rr_max_c = col_bounds
                rr_content_bounds = (content_bounds[0], content_bounds[1], rr_min_c, rr_max_c)
            else:
                rr_content_bounds = content_bounds

            files = _export_sheet_screenshot(
                ws_snap, output_dir, sheet_name,
                content_bounds=rr_content_bounds,
                row_range=(start_row, end_row),
                output_filename=filename
            )
            results.append({
                'filename': filename,
                'start_row': start_row,
                'end_row': end_row,
                'success': len(files) > 0,
                'files': files,
                'error': None if files else '截图生成失败',
            })
        except Exception as e:
            results.append({
                'filename': filename,
                'start_row': start_row,
                'end_row': end_row,
                'success': False,
                'files': [],
                'error': str(e),
            })
        finally:
            # 每个 zone 截图后立即销毁 COM 实例
            try:
                if wb_snap:
                    wb_snap.Close(SaveChanges=False)
            except:
                pass
            try:
                if excel_snap:
                    excel_snap.Quit()
            except:
                pass
            wb_snap = None
            excel_snap = None

    pythoncom.CoUninitialize()
    return results


# ──────────────────────────────────────────────
# 命令行入口
# ──────────────────────────────────────────────

def main():
    import argparse as _argparse
    parser = _argparse.ArgumentParser(description="Excel 双通道导出工具 (截图 + 结构化文本)")
    parser.add_argument("input_path", help="输入 Excel 文件路径")
    parser.add_argument("output_dir", nargs="?", default=None, help="输出目录（默认: 同目录下 {tag}_export/）")
    parser.add_argument("--sheets", nargs="+", default=None, help="要导出的 sheet 名称列表（空格分隔），默认导出全部")
    parser.add_argument("--tag", default=None, help="输出目录标签（ASCII），拼接为 {tag}_export/")

    args = parser.parse_args()

    try:
        export_excel(args.input_path, args.output_dir, selected_sheets=args.sheets, tag=args.tag)
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
